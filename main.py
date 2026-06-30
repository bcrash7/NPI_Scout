#!/usr/bin/env python3
"""
concierge_npi_scout.py
Look up a physician by NPI and open a self-contained HTML report in Chrome.
Usage: python main.py 1234567890 [--json] [--no-google] [--no-scrape]
"""

from __future__ import annotations

import argparse
import base64
import datetime as dt
import json
import math
import os
import re
import sys
import tempfile
import time
import urllib.parse
import webbrowser
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any, Optional

import requests

try:
    from bs4 import BeautifulSoup
    _HAVE_BS4 = True
except ImportError:
    _HAVE_BS4 = False

try:
    import openpyxl
    _HAVE_OPENPYXL = True
except ImportError:
    _HAVE_OPENPYXL = False


# ----------------------------------------------------------------------------
# Configuration
# ----------------------------------------------------------------------------
GOOGLE_API_KEY = os.environ.get("GOOGLE_MAPS_API_KEY", "").strip()
CENSUS_API_KEY = os.environ.get("CENSUS_API_KEY", "").strip()
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "").strip()
# Optional: a Google Programmable Search Engine ID ("cx"). When set alongside the
# Google key, directory profiles (Healthgrades/US News/etc.) are discovered via the
# Google Custom Search JSON API. Without it, discovery falls back to DuckDuckGo (no key).
GOOGLE_CSE_ID = os.environ.get("GOOGLE_CSE_ID", "").strip()
# Experian Mosaic by-ZIP workbook (counts per Mosaic code) used for the consumer-fit
# score. Defaults to a file next to this script; override with MOSAIC_XLSX.
MOSAIC_XLSX = os.environ.get(
    "MOSAIC_XLSX", str(Path(__file__).resolve().parent / "ZIP_Mosaic_Pop_Added.xlsx"))
# monday.com push (optional). Token from monday Admin -> API; board/group from the
# board URL. Columns are matched to the board by TITLE, so just name columns to match
# (e.g. "Fit Score", "Lane", "Mosaic Score") — no need to copy column IDs.
MONDAY_API_TOKEN = os.environ.get("MONDAY_API_TOKEN", "").strip()
MONDAY_BOARD_ID = os.environ.get("MONDAY_BOARD_ID", "").strip()
MONDAY_GROUP_ID = os.environ.get("MONDAY_GROUP_ID", "").strip()
MONDAY_API_URL = "https://api.monday.com/v2"
MONDAY_API_VERSION = "2024-04"
# Primary model for reasoning/extraction from page text. Override with NPI_SCOUT_MODEL
# (e.g. "claude-opus-4-8").
LLM_MODEL = os.environ.get("NPI_SCOUT_MODEL", "claude-sonnet-4-6").strip()
# Token-heavy / high-volume calls (open web search, which pulls whole pages into context, and
# image/photo checks) run on a lighter, higher-throughput model BY DEFAULT. Rate limits are
# per-model, so this keeps those calls from exhausting the primary model's per-minute token
# budget (the website-analysis call then runs uncontended on the primary model). Override with
# NPI_SCOUT_AUX_MODEL; set it equal to NPI_SCOUT_MODEL to force everything onto one model.
LLM_MODEL_AUX = os.environ.get("NPI_SCOUT_AUX_MODEL", "claude-haiku-4-5-20251001").strip()
print(f"[config] Google key: {'detected' if GOOGLE_API_KEY else 'MISSING'}; "
      f"Census key: {'detected' if CENSUS_API_KEY else 'MISSING'}; "
      f"Anthropic key: {'detected' if ANTHROPIC_API_KEY else 'MISSING'}; "
      f"CSE id: {'detected' if GOOGLE_CSE_ID else 'using DuckDuckGo'}", file=sys.stderr)

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36 ConciergeNPIScout/1.3"
)

NPPES_URL = "https://npiregistry.cms.hhs.gov/api/"
CMS_DATA_JSON = "https://data.cms.gov/data.json"
CMS_DATA_API = "https://data.cms.gov/data-api/v1/dataset/{uuid}/data"
# CMS "Doctors and Clinicians" National Downloadable File (Provider Data Catalog,
# DKAN datastore). datasetID is stable across refreshes; distribution index is always 0.
# Source of medical school, graduation year, and Medicare primary/secondary specialties.
DAC_DATASET_ID = "mj5m-pzi6"
DAC_QUERY_URL = "https://data.cms.gov/provider-data/api/1/datastore/query/{ds}/0"
CENSUS_BASE = "https://api.census.gov/data/{year}/acs/acs5"
CENSUS_GEOCODER = "https://geocoding.geo.census.gov/geocoder/locations/onelineaddress"
CENSUS_YEARS_TO_TRY = [2023, 2022, 2021]

OSM_STATICMAP = "https://staticmap.openstreetmap.de/staticmap.php"

# Physician directories to look up for star ratings / review counts and profile links.
# (label, domain, path-keywords that indicate a real provider profile vs a listing page)
AGGREGATOR_SITES = [
    ("Healthgrades", "healthgrades.com", ("physician", "providers", "dr-")),
    ("U.S. News", "health.usnews.com", ("doctors",)),
    ("Vitals", "vitals.com", ("doctor", "dr_", "dr-")),
    ("WebMD", "webmd.com", ("doctor", "physician")),
    ("Zocdoc", "zocdoc.com", ("doctor",)),
]

# ----------------------------------------------------------------------------
# Concierge Fit scoring rubric  (edit these to retune; all are explicit knobs)
# ----------------------------------------------------------------------------
# Primary-specialty fit. Matching is done on normalized substrings against the
# NPPES taxonomy description and the CMS primary specialty.
SPECIALTY_HIGH = [
    "internal medicine", "family medicine", "family practice",
    "geriatric", "general practice",
]
SPECIALTY_OK = [
    "endocrinolog", "cardiolog", "cardiovascular", "gastroenterolog", "rheumatolog",
    "pulmonolog", "pulmonary", "nephrolog", "allergy", "immunolog", "infectious disease",
]
# A primary specialty in neither list scores 0 AND hard-disqualifies (concierge
# conversion is specialty-gated). Flip SPECIALTY_OTHER_IS_DEAD to False to instead
# only zero the specialty component without auto-DEAD.
SPECIALTY_OTHER_IS_DEAD = True

# Employment detection. A billing/affiliation org name matching these patterns is
# treated as hospital/health-system EMPLOYMENT -> automatic DEAD (score 0).
# Pure affiliation (admitting privileges, IPA membership) is NOT employment; this
# only fires on the entity the clinician bills under or a website employment claim.
EMPLOYER_PATTERNS = [
    "hospital", "health system", "healthcare system", "health network",
    "medical center", "med center", "medical foundation", "health foundation",
    "university", "kaiser", "ascension", "providence health", "sutter",
    "geisinger", "cleveland clinic", "mayo clinic", "intermountain", "advocate",
    "trinity health", "commonspirit", "common spirit", "tenet", "hca healthcare",
    "baptist health", "veterans affairs", "va medical", "permanente",
]

# Career-stage curve, centered on the average affiliate age. Plateau scores 100;
# younger doctors score progressively lower; 75+ tapers slightly.
AFFILIATE_AVG_AGE = 61
CAREER_PLATEAU_LOW, CAREER_PLATEAU_HIGH = 55, 74
CAREER_YOUNG_SLOPE = 4.0   # points lost per year below the plateau
CAREER_OLD_SLOPE = 3.0     # points lost per year above 74 (slight)

# Composite weights (must sum to 100). Payer mix excluded. Independence/ownership is
# NOT scored — it is surfaced as flags instead (see compute_fit_score): a group of 6+
# providers is a yellow flag; hospital/health-system employment is a red flag.
FIT_WEIGHTS = {
    "affluence": 40,        # Census area income blended with Mosaic Score
    "medicare_volume": 25,  # absorbed the old independence weight; 400+ benes scores great
    "career_stage": 25,
    "specialty": 10,
}
# Final-score -> triage lane (hard disqualifiers override to DEAD).
FIT_LANES = [("HOT", 70), ("WARM", 50), ("LATER", 30), ("DEAD", 0)]

# ---- Market context: concierge competition & affiliate proximity ----
# Concierge competitors near a practice. DPC (direct primary care) is intentionally
# NOT counted as competition. Matching is on the listing name.
COMPETITION_QUERIES = ["concierge medicine", "concierge doctor", "membership medicine"]
COMPETITION_KEYWORDS = ["concierge", "mdvip", "specialdocs", "partnermd",
                        "castle connolly", "concierge choice", "private health"]
COMPETITION_EXCLUDE = ["direct primary care", "dpc", "signaturemd", "signature md"]
COMPETITION_RADIUS_MI = 10
# National concierge/management/franchise brands whose org NPI lists locations all over
# the country — these are NOT the physician's own offices, so we never import their
# location lists, and we keep only offices near the doctor's primary practice address.
NATIONAL_BRANDS = ["signaturemd", "signature md", "mdvip", "specialdocs", "partnermd",
                   "castle connolly", "concierge choice", "privia", "optum", "oak street",
                   "one medical", "village md", "villagemd"]
# Offices farther than this from the primary practice address are treated as historical
# / network addresses and set aside (not counted as current offices).
OFFICE_LOCALITY_MI = 60
# Affiliate locations CSV (optional): needs a name column and latitude/longitude columns.
# Distance to the nearest affiliate is reported as a warm-intro signal.
AFFILIATES_CSV = os.environ.get(
    "AFFILIATES_CSV", str(Path(__file__).resolve().parent / "AffiliateGeo.csv"))
AFFILIATE_NEARBY_MI = 50  # count affiliates within this radius

# Mid-level providers to count: ONLY nurse practitioners and physician assistants.
# Credentials are matched as prefixes, so suffixes like -BC / -C are tolerated
# (e.g. "FNP-BC" -> NP, "PA-C" -> PA). Anyone not matching (CNM, CRNA, CNS, RN, MA,
# physicians) is excluded from the midlevel count.
NP_CREDENTIALS = ["APRN", "FNP", "PMHNP", "AGPCNP", "AGACNP", "WHNP", "PNP",
                  "NNP", "ENP", "DNP", "NP"]
PA_CREDENTIALS = ["PA-C", "DMSC", "DHSC", "PA"]


_STATE_FIPS = {
    "AL": "01", "AK": "02", "AZ": "04", "AR": "05", "CA": "06", "CO": "08",
    "CT": "09", "DE": "10", "DC": "11", "FL": "12", "GA": "13", "HI": "15",
    "ID": "16", "IL": "17", "IN": "18", "IA": "19", "KS": "20", "KY": "21",
    "LA": "22", "ME": "23", "MD": "24", "MA": "25", "MI": "26", "MN": "27",
    "MS": "28", "MO": "29", "MT": "30", "NE": "31", "NV": "32", "NH": "33",
    "NJ": "34", "NM": "35", "NY": "36", "NC": "37", "ND": "38", "OH": "39",
    "OK": "40", "OR": "41", "PA": "42", "RI": "44", "SC": "45", "SD": "46",
    "TN": "47", "TX": "48", "UT": "49", "VT": "50", "VA": "51", "WA": "53",
    "WV": "54", "WI": "55", "WY": "56", "PR": "72",
}

INSURANCE_KEYWORDS = [
    "medicare", "medicaid", "aetna", "cigna", "humana", "united healthcare",
    "unitedhealthcare", "blue cross", "blue shield", "bcbs", "anthem", "tricare",
    "kaiser", "oscar", "molina", "wellcare", "centene", "self-pay", "self pay",
    "out of network", "out-of-network", "ppo", "hmo", "epo", "we accept most",
]
SERVICE_KEYWORDS = [
    # Primary care / concierge internal medicine
    "annual physical", "wellness exam", "preventive care", "preventative care",
    "chronic disease", "chronic care", "diabetes", "hypertension", "cardiology",
    "telemedicine", "telehealth", "same-day", "same day", "urgent care",
    "lab work", "bloodwork", "blood work", "vaccinations", "immunizations",
    "weight management", "weight loss", "hormone", "iv therapy", "concierge",
    "membership", "house call", "house calls", "geriatric", "pediatric",
    "women's health", "men's health", "physical therapy", "dermatology",
    "screening", "ekg", "ecg", "spirometry", "ultrasound",
    # Aesthetic / med-spa (so a med spa is recognizable, not mislabeled as primary care)
    "med spa", "medspa", "medical spa", "aesthetic", "aesthetics", "esthetician",
    "botox", "dysport", "jeuveau", "neurotoxin", "botulinum", "dermal filler",
    "filler", "lip filler", "injectable", "injectables", "microneedling",
    "chemical peel", "facial", "facials", "hydrafacial", "dermaplaning",
    "laser hair removal", "laser", "ipl", "photofacial", "coolsculpting",
    "body contouring", "skin tightening", "morpheus8", "prp",
    "platelet-rich plasma", "microblading", "skin rejuvenation", "anti-aging",
    "anti aging", "wrinkle", "peels", "hormone replacement", "testosterone",
    "trt", "semaglutide", "tirzepatide", "vitamin injection", "b12 injection",
]

_CERT_PATTERNS = [
    r"board[\s-]*certified(?:\s+in\s+[A-Za-z ,&/]+)?",
    r"diplomate[\s,]+(?:of\s+)?the\s+american\s+board\s+of\s+[A-Za-z ,&/]+",
    r"american\s+board\s+of\s+[A-Za-z ,&/]+",
    r"fellow\s+of\s+the\s+american\s+(?:college|academy)\s+of\s+[A-Za-z ,&/]+",
]
_RECOGNITION_PATTERNS = [
    r"castle\s+connolly", r"top\s+doctor[s]?", r"best\s+doctor[s]?",
    r"patients?'?\s+choice\s+award", r"america'?s?\s+top\s+\w+",
]


# ----------------------------------------------------------------------------
# HTTP helpers
# ----------------------------------------------------------------------------
_SESSION = requests.Session()
_SESSION.headers.update({"User-Agent": USER_AGENT})


def _get(url, *, params=None, headers=None, timeout=25, retries=3, expect="json", quiet=False):
    last_err = None
    for attempt in range(retries):
        try:
            r = _SESSION.get(url, params=params, headers=headers, timeout=timeout)
            if r.status_code == 429:
                time.sleep(2 * (attempt + 1))
                continue
            r.raise_for_status()
            if expect == "json":
                return r.json()
            if expect == "text":
                return r.text
            if expect == "bytes":
                return r.content
            return r
        except Exception as e:  # noqa: BLE001
            last_err = e
            if attempt < retries - 1:
                time.sleep(1.0 * (attempt + 1))
    if last_err and not quiet:
        print(f"    [warn] GET failed ({url.split('?')[0]}): {last_err}", file=sys.stderr)
    return None


def _post(url, *, json_body, headers=None, timeout=25, retries=5):
    last_err = None
    last_body = None
    for attempt in range(retries):
        try:
            r = _SESSION.post(url, json=json_body, headers=headers, timeout=timeout)
            # Rate limited / overloaded: wait out the window and retry. Honor Retry-After
            # (seconds) when the server sends it; otherwise back off exponentially. This is
            # what keeps low input-token-per-minute tiers from dropping calls.
            if r.status_code in (429, 529) and attempt < retries - 1:
                ra = r.headers.get("retry-after") or r.headers.get("Retry-After")
                try:
                    wait = float(ra) if ra is not None else 0.0
                except ValueError:
                    wait = 0.0
                if wait <= 0:
                    wait = 5.0 * (2 ** attempt)         # 5, 10, 20, 40s
                wait = min(wait + 1.0, 60.0)             # small cushion; cap at the 1-min window
                print(f"    [info] {r.status_code} (rate limit) from {url.split('?')[0]}; waiting "
                      f"{wait:.0f}s then retrying ({attempt + 1}/{retries}).", file=sys.stderr)
                time.sleep(wait)
                continue
            if r.status_code >= 400:
                last_body = r.text
            r.raise_for_status()
            return r.json()
        except Exception as e:  # noqa: BLE001
            last_err = e
            if attempt < retries - 1:
                time.sleep(1.0 * (attempt + 1))
    if last_err:
        print(f"    [warn] POST failed ({url.split('?')[0]}): {last_err}", file=sys.stderr)
        if last_body:
            print(f"    [warn] response body: {last_body[:600]}", file=sys.stderr)
    return None


def fetch_image_b64(url, *, headers=None, max_bytes=5_000_000) -> Optional[str]:
    r = _get(url, headers=headers, expect="resp", timeout=25, retries=2)
    if r is None:
        return None
    ctype = (r.headers.get("Content-Type", "") or "").split(";")[0].strip().lower()
    if not ctype.startswith("image/"):
        return None
    data = r.content
    if not data or len(data) > max_bytes:
        return None
    return f"data:{ctype};base64,{base64.b64encode(data).decode('ascii')}"


# ----------------------------------------------------------------------------
# Data model
# ----------------------------------------------------------------------------
@dataclass
class Address:
    purpose: str = ""
    line1: str = ""
    line2: str = ""
    city: str = ""
    state: str = ""
    postal: str = ""
    phone: str = ""
    lat: Optional[float] = None
    lng: Optional[float] = None
    area_median_household_income: Optional[int] = None
    area_per_capita_income: Optional[int] = None
    area_median_home_value: Optional[int] = None
    census_year: Optional[int] = None
    photo_data_uris: list[str] = field(default_factory=list)
    street_view_data_uri: str = ""
    satellite_data_uri: str = ""
    map_data_uri: str = ""
    map_link: str = ""
    source: str = "NPPES (individual)"
    mosaic_score: Optional[float] = None
    mosaic_high_value_share: Optional[float] = None
    mosaic_classified_pop: Optional[int] = None
    mosaic_top5: list = field(default_factory=list)  # [(code, name, segment, count, share, lh)]
    superseded: bool = False   # an NPPES address replaced by a current website/listing address

    @property
    def zip5(self) -> str:
        m = re.match(r"\d{5}", self.postal or "")
        return m.group(0) if m else ""

    @property
    def is_office(self) -> bool:
        return (not self.superseded) and self.purpose in ("LOCATION", "PRACTICE LOCATION")

    def one_line(self) -> str:
        bits = [self.line1, self.line2, f"{self.city}, {self.state} {self.postal}".strip()]
        return ", ".join(b for b in bits if b and b.strip(", "))


@dataclass
class ReviewSource:
    source: str = ""
    rating: Optional[float] = None
    review_count: Optional[int] = None
    url: str = ""
    note: str = ""


@dataclass
class ProviderProfile:
    npi: str = ""
    retrieved_at: str = ""
    first_name: str = ""
    last_name: str = ""
    middle_name: str = ""
    credential: str = ""
    gender: str = ""
    sole_proprietor: str = ""
    organization_name: str = ""
    practice_name: str = ""
    affiliated_org_name: str = ""
    places_name: str = ""
    enumeration_type: str = ""
    enumeration_date: str = ""
    last_updated: str = ""
    years_in_practice_proxy: Optional[float] = None
    estimated_age: Optional[int] = None
    age_basis: str = ""
    medical_school: str = ""
    medical_school_source: str = ""
    residency: str = ""
    residency_source: str = ""
    tenure_at_location: str = ""
    tenure_source: str = ""
    graduation_year: Optional[int] = None
    specializations: list[str] = field(default_factory=list)
    primary_taxonomy: str = ""
    dac_primary_specialty: str = ""
    dac_secondary_specialties: list[str] = field(default_factory=list)
    licenses: list[str] = field(default_factory=list)
    addresses: list[Address] = field(default_factory=list)
    medicare_beneficiaries: Optional[int] = None
    medicare_total_services: Optional[float] = None
    medicare_total_payment: Optional[float] = None
    medicare_drug_beneficiaries: Optional[int] = None
    medicare_medical_beneficiaries: Optional[int] = None
    medicare_data_year: Optional[int] = None
    medicare_provider_type: str = ""
    website: str = ""
    practice_type: str = ""
    practice_summary: str = ""
    website_offices: list = field(default_factory=list)   # [dict(line1,line2,city,state,postal,phone)]
    places_formatted_address: str = ""                    # Google Business listing address (fallback)
    web_offices: list = field(default_factory=list)       # current office(s) found via open web search
    web_current_employer: str = ""                        # current employer/group per open web search
    web_evidence: list = field(default_factory=list)      # source URLs the web search relied on
    web_confidence: str = ""                              # high / medium / low
    personal_interests: list = field(default_factory=list)  # [{"text":.., "source":..}] non-clinical rapport facts
    website_services: list[str] = field(default_factory=list)
    website_insurances: list[str] = field(default_factory=list)
    reviews: list[ReviewSource] = field(default_factory=list)
    external_profiles: list[ReviewSource] = field(default_factory=list)
    board_certifications: list[str] = field(default_factory=list)
    recognitions: list[str] = field(default_factory=list)
    accepts_insurance: Optional[bool] = None
    interior_photos: list[str] = field(default_factory=list)      # data URIs
    provider_photos: list = field(default_factory=list)           # [(label, data_uri)]
    midlevels: list = field(default_factory=list)                 # [(name, type)]
    midlevel_count: Optional[int] = None                          # None => not verifiable
    midlevel_roster_known: bool = False
    # Conversion signals
    group_size: Optional[int] = None          # CMS num_org_mem
    dac_org_name: str = ""                     # CMS Facility Name (billing org)
    employment_status: str = ""                # independent / likely_employed / unknown
    employment_evidence: str = ""
    career_stage: str = ""
    panel_estimate: Optional[int] = None       # informational only (not scored)
    specialty_fit_tier: str = ""               # high / acceptable / none
    npi_deactivated: bool = False
    acquisition_target: bool = False
    acquisition_reason: str = ""
    flags: list = field(default_factory=list)  # [(level 'red'|'yellow', text)]
    # Market context (informational, not folded into the Fit Score)
    competitor_count: Optional[int] = None
    competitors: list = field(default_factory=list)  # [(name, miles)]
    nearest_affiliate_name: str = ""
    nearest_affiliate_mi: Optional[float] = None
    affiliates_nearby: Optional[int] = None
    # Concierge Fit Score
    fit_score: Optional[int] = None
    fit_lane: str = ""
    fit_components: list = field(default_factory=list)  # [(label, weight, subscore, weighted)]
    disqualifiers: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    sources_used: list[str] = field(default_factory=list)

    @property
    def full_name(self) -> str:
        return " ".join(b for b in [self.first_name, self.last_name] if b) or self.organization_name

    @property
    def display_practice_name(self) -> str:
        """Best available practice name, in priority order. Skips any candidate that
        is just the doctor's own name (e.g. a Google listing titled 'Sohan Varma, MD')."""
        doc = self.full_name.strip().lower()
        for c in (self.practice_name, self.affiliated_org_name,
                  self.organization_name, self.places_name):
            c = (c or "").strip()
            if c and c.lower() != doc and c.lower() not in doc and doc not in c.lower():
                return c
        return ""

    @property
    def additional_specializations(self) -> list[str]:
        """NPPES taxonomies other than the primary one, cleaned of the '(primary)' tag."""
        out = []
        for s in self.specializations:
            label = s.replace(" (primary)", "").strip()
            if "(primary)" in s:
                continue
            if label and label != self.primary_taxonomy and label not in out:
                out.append(label)
        return out


# ----------------------------------------------------------------------------
# NPPES
# ----------------------------------------------------------------------------
def fetch_nppes(npi: str, prof: ProviderProfile) -> None:
    data = _get(NPPES_URL, params={"version": "2.1", "number": npi})
    if not data or not data.get("results"):
        prof.notes.append("NPPES: no record found for this NPI.")
        return
    prof.sources_used.append("NPPES NPI Registry")
    rec = data["results"][0]
    prof.enumeration_type = rec.get("enumeration_type", "")
    basic = rec.get("basic", {})
    prof.first_name = basic.get("first_name", "")
    prof.last_name = basic.get("last_name", "")
    prof.middle_name = basic.get("middle_name", "")
    prof.credential = basic.get("credential", "")
    prof.gender = basic.get("gender", "")
    prof.sole_proprietor = basic.get("sole_proprietor", "")
    prof.organization_name = basic.get("organization_name", "") or basic.get("name", "")
    prof.enumeration_date = basic.get("enumeration_date", "")
    prof.last_updated = basic.get("last_updated", "")
    if (basic.get("status", "") or "").upper() == "D" or basic.get("deactivation_date"):
        prof.npi_deactivated = True
        prof.notes.append("NPPES record is deactivated"
                          + (f" (date {basic.get('deactivation_date')})" if basic.get("deactivation_date") else "")
                          + ".")

    if prof.enumeration_date:
        try:
            ed = dt.datetime.strptime(prof.enumeration_date, "%Y-%m-%d").date()
            prof.years_in_practice_proxy = round((dt.date.today() - ed).days / 365.25, 1)
        except ValueError:
            pass

    for tax in rec.get("taxonomies", []):
        desc = tax.get("desc", "")
        if desc:
            prof.specializations.append(desc + (" (primary)" if tax.get("primary") else ""))
            if tax.get("primary"):
                prof.primary_taxonomy = desc
        lic, st = tax.get("license", ""), tax.get("state", "")
        if lic:
            prof.licenses.append(f"{lic} ({st})" if st else lic)
    if not prof.primary_taxonomy and prof.specializations:
        prof.primary_taxonomy = prof.specializations[0].replace(" (primary)", "")

    for ad in rec.get("addresses", []):
        prof.addresses.append(Address(
            purpose=ad.get("address_purpose", ""), line1=ad.get("address_1", ""),
            line2=ad.get("address_2", ""), city=ad.get("city", ""),
            state=ad.get("state", ""), postal=ad.get("postal_code", ""),
            phone=ad.get("telephone_number", ""),
        ))
    for ad in rec.get("practiceLocations", []) or []:
        prof.addresses.append(Address(
            purpose="PRACTICE LOCATION", line1=ad.get("address_1", ""),
            line2=ad.get("address_2", ""), city=ad.get("city", ""),
            state=ad.get("state", ""), postal=ad.get("postal_code", ""),
            phone=ad.get("telephone_number", ""),
        ))


def _addr_key(line1: str, postal: str) -> str:
    """Normalized key for de-duplicating addresses across NPPES records."""
    l = re.sub(r"\s+", " ", (line1 or "")).strip().lower()
    m = re.match(r"\d{5}", postal or "")
    z = m.group(0) if m else ""
    return f"{l}|{z}"


# Street-word and directional normalization so "490 Post Street" == "490 Post St" and
# "Ste 900" == "# 900" == "Suite 900". All unit designators collapse to a single token so
# only the unit *number* distinguishes suites.
_STREET_TOKEN_MAP = {
    "street": "st", "st": "st", "avenue": "ave", "ave": "ave", "av": "ave",
    "boulevard": "blvd", "blvd": "blvd", "road": "rd", "rd": "rd", "drive": "dr", "dr": "dr",
    "lane": "ln", "ln": "ln", "court": "ct", "ct": "ct", "place": "pl", "pl": "pl",
    "parkway": "pkwy", "pkwy": "pkwy", "highway": "hwy", "hwy": "hwy", "terrace": "ter",
    "ter": "ter", "circle": "cir", "cir": "cir", "square": "sq", "sq": "sq",
    "trail": "trl", "trl": "trl", "way": "way", "plaza": "plz", "plz": "plz",
    "north": "n", "south": "s", "east": "e", "west": "w",
    "northeast": "ne", "northwest": "nw", "southeast": "se", "southwest": "sw",
}
_UNIT_MARKERS = {"ste", "suite", "unit", "apt", "apartment", "rm", "room", "fl", "floor",
                 "flr", "bldg", "building", "no", "num", "number", "lobby", "lbby",
                 "dept", "department", "#"}


def _canon_addr(line1: str, line2: str = ""):
    """Return (strict, loose, has_unit). 'strict' includes the unit designator collapsed to a
    generic 'u' token plus its value; 'loose' is the street portion only (no unit). Used to
    recognize the same location across formatting/mailing-vs-practice variants."""
    raw = f"{line1 or ''} {line2 or ''}".lower()
    raw = raw.replace("#", " # ")
    raw = re.sub(r"[.,]", " ", raw)
    toks = [t for t in re.split(r"\s+", raw) if t]
    full, has_unit = [], False
    for t in toks:
        if t in _UNIT_MARKERS:
            full.append("u")
            has_unit = True
        else:
            full.append(_STREET_TOKEN_MAP.get(t, t))
    strict = " ".join(full).strip()
    loose = []
    for t in full:
        if t == "u":
            break
        loose.append(t)
    return strict, " ".join(loose).strip(), has_unit


def _addr_rank(a) -> tuple:
    """Higher is better: prefer an office over a mailing copy, with a phone, geocoded, and
    the most complete one-line text."""
    return (1 if a.is_office else 0, 1 if a.phone else 0,
            1 if a.lat is not None else 0, len(a.one_line()))


def dedupe_addresses(prof: ProviderProfile) -> None:
    """Collapse addresses that are the same place under different formatting — St vs Street,
    Ste vs #, ZIP+4 vs ZIP5, or the same office listed as both MAILING and LOCATION and again
    by an affiliated-org NPI. Without this, one physical office can render as several 'offices'.
    The richest record per location is kept; its source note records the merged listings."""
    if len(prof.addresses) <= 1:
        return
    meta = {}
    for a in prof.addresses:
        strict, loose, has_unit = _canon_addr(a.line1, a.line2)
        meta[id(a)] = (strict, loose, has_unit, a.zip5)

    # Pass 1: exact canonical street (incl. unit) + ZIP5.
    groups, order = {}, []
    for a in prof.addresses:
        strict, _, _, z = meta[id(a)]
        skey = f"{strict}|{z}"
        if skey not in groups:
            groups[skey] = []
            order.append(skey)
        groups[skey].append(a)
    survivors, collapsed = [], 0
    for skey in order:
        g = groups[skey]
        best = max(g, key=_addr_rank)
        others = [x.source for x in g if x is not best and getattr(x, "source", "")]
        extra = "; ".join(s for s in dict.fromkeys(others) if s and s not in (best.source or ""))
        if extra:
            best.source = (best.source + "; also listed by " + extra) if best.source else extra
        survivors.append(best)
        collapsed += len(g) - 1

    # Pass 2: a street-only listing (no unit) at the same street+ZIP as a unit-bearing listing
    # is the less-specific duplicate of it -> drop the street-only one.
    by_loose = {}
    for a in survivors:
        _, loose, _, z = meta[id(a)]
        by_loose.setdefault(f"{loose}|{z}", []).append(a)
    drop = set()
    for g in by_loose.values():
        if len(g) >= 2 and any(meta[id(x)][2] for x in g):
            for x in g:
                if not meta[id(x)][2]:
                    drop.add(id(x))
                    collapsed += 1
    survivors = [a for a in survivors if id(a) not in drop]

    if collapsed:
        prof.addresses = survivors
        prof.notes.append(f"Merged {collapsed} duplicate address listing(s) for the same "
                          "location (formatting / mailing-vs-practice / affiliated-org variants).")


def collapse_colocated_offices(prof: ProviderProfile) -> None:
    """Backstop after geocoding: office rows that resolve to essentially the same point and ZIP
    are one place under formatting the text normalizer missed. Keep the richer row."""
    offices = [a for a in prof.addresses
               if a.is_office and a.lat is not None and a.lng is not None]
    if len(offices) <= 1:
        return
    drop = set()
    for i in range(len(offices)):
        for j in range(i + 1, len(offices)):
            a, b = offices[i], offices[j]
            if id(a) in drop or id(b) in drop:
                continue
            if a.zip5 and b.zip5 and a.zip5 != b.zip5:
                continue
            if _haversine_mi(a.lat, a.lng, b.lat, b.lng) <= 0.1:  # ~150 m
                drop.add(id(a if _addr_rank(a) <= _addr_rank(b) else b))
    if drop:
        prof.addresses = [a for a in prof.addresses if id(a) not in drop]
        prof.notes.append(f"Merged {len(drop)} office row(s) that geocoded to the same location.")


def _parse_formatted_us_address(s: str) -> dict:
    """Loosely split a Google 'formattedAddress' (e.g. '2433 Country Place Blvd Bldg B,
    Trinity, FL 34655, USA') into components for use as a current-address fallback."""
    s = re.sub(r",?\s*USA\s*$", "", (s or "").strip())
    parts = [p.strip() for p in s.split(",") if p.strip()]
    if len(parts) < 2:
        return {}
    line1 = parts[0]
    city = state = postal = ""
    m = re.search(r"\b([A-Z]{2})\s+(\d{5}(?:-\d{4})?)\b", parts[-1])
    if m:
        state, postal = m.group(1), m.group(2)
        city = parts[-2] if len(parts) >= 3 else ""
    elif re.fullmatch(r"[A-Z]{2}", parts[-1]):
        state = parts[-1]
        city = parts[-2] if len(parts) >= 2 else ""
    else:
        city = parts[1]
    if not line1 or not (city or postal):
        return {}
    return {"line1": line1, "line2": "", "city": city, "state": state,
            "postal": postal, "phone": ""}


def apply_current_address(prof: ProviderProfile) -> None:
    """Make the practice's current address authoritative over NPPES. NPPES address data is
    frequently years out of date, so the current office(s) are taken (in priority order) from
    the provider's own website, an open web search, or the Google Business listing. NPPES office
    addresses are superseded — kept for reference and flagged, but no longer enriched/mapped."""
    sourced = []  # (components_dict, source_label)
    for comp in (prof.website_offices or []):
        sourced.append((comp, "Provider website (scraped)"))
    if not sourced:
        for comp in (prof.web_offices or []):
            sourced.append((comp, "Open web search"))
    if not sourced and prof.places_formatted_address:
        g = _parse_formatted_us_address(prof.places_formatted_address)
        if g:
            sourced.append((g, "Google Business listing"))
    if not sourced:
        return

    new_offices = []
    for comp, label in sourced:
        new_offices.append(Address(
            purpose="LOCATION", line1=comp["line1"], line2=comp.get("line2", ""),
            city=comp["city"], state=comp["state"], postal=comp["postal"],
            phone=comp.get("phone", ""), source=label))

    # If the current address simply confirms an NPPES office we already have, mark that office
    # confirmed (and promote it) rather than superseding everything.
    existing_office_keys = {(_canon_addr(a.line1, a.line2)[0], a.zip5): a
                            for a in prof.addresses if a.is_office}
    confirmed_only = (len(new_offices) == 1 and
                      (_canon_addr(new_offices[0].line1, new_offices[0].line2)[0],
                       new_offices[0].zip5) in existing_office_keys)
    if confirmed_only:
        a = existing_office_keys[(_canon_addr(new_offices[0].line1, new_offices[0].line2)[0],
                                  new_offices[0].zip5)]
        label = new_offices[0].source.split(" (")[0]
        if label.lower() not in (a.source or "").lower():
            a.source = (a.source + f"; confirmed current via {label}") if a.source else new_offices[0].source
        prof.addresses.remove(a)
        prof.addresses.insert(0, a)
        return

    superseded = 0
    for a in prof.addresses:
        if a.is_office:
            a.superseded = True
            superseded += 1
    prof.addresses = new_offices + prof.addresses
    src = new_offices[0].source
    note = f"Address taken from the {src.lower()} as the current source; "
    if superseded:
        note += (f"{superseded} NPPES-registered office address(es) were superseded "
                 "(kept for reference below, but NPPES address data often lags the practice's "
                 "current location).")
    else:
        note += "no NPPES office address was on file."
    prof.notes.append(note)


def find_current_info_via_web(prof: ProviderProfile, want_location: bool = True,
                              want_interests: bool = True) -> None:
    """One open web search (Anthropic server-side web_search tool) that, in a single pass over the
    physician's public pages, can establish (a) WHERE they practice NOW — NPPES is frequently
    years out of date — and (b) PUBLIC, non-clinical rapport facts (hobbies, sports, music,
    volunteering). Doing both in one call keeps web-search token usage (whole pages are pulled
    into context) within tight per-minute limits, and it runs on LLM_MODEL_AUX so it doesn't
    consume the primary model's budget that the website-analysis call needs."""
    if not ANTHROPIC_API_KEY or not (want_location or want_interests):
        return
    name = prof.full_name
    spec = prof.primary_taxonomy or prof.dac_primary_specialty or ""
    org = (prof.web_current_employer or prof.dac_org_name or prof.affiliated_org_name
           or prof.organization_name or "").strip()
    o = _primary_office(prof)
    nppes_line = o.one_line() if o else "unknown"

    parts = [
        "You are researching one specific U.S. physician. Use web search to find the most current, "
        "authoritative PUBLIC information about THIS exact person (match the NPI, full name, and "
        "specialty; ignore same-named others).",
        f"- Name: {name or 'UNKNOWN — identify the provider from the NPI'}",
        f"- NPI: {prof.npi}", f"- Specialty: {spec or 'unknown'}",
        f"- Employer/group on file (CMS, may help): {org or 'unknown'}",
        f"- Address on file with NPPES (OFTEN OUT OF DATE — do not trust it): {nppes_line}",
        "",
    ]
    schema = []
    if want_location:
        parts.append(
            "TASK A — IDENTITY & CURRENT PRACTICE LOCATION: confirm the provider's full name (look "
            "it up from the NPI if it is unknown above) and find where this physician practices "
            "NOW. Prefer the employer/health-system official 'find a doctor' profile or the "
            "practice's own site; ignore outdated listings. Never guess an address.")
        schema.append(
            '"provider_name": <full name or null>, "provider_credential": <e.g. MD/DO or null>, '
            '"current_employer": <string or null>, "website": <official current practice/profile '
            'URL or null>, "offices": [{"line1":..,"line2":.. or null,"city":..,"state":<2-letter>,'
            '"postal":..,"phone":.. or null}], "confidence": "high"|"medium"|"low", '
            '"evidence_urls": [<urls>]')
    if want_interests:
        parts.append(
            "TASK B — PERSONAL INTERESTS (for rapport): find PUBLIC, non-clinical human-interest "
            "facts — hobbies, sports (marathons, cycling), music (plays in a band), arts/crafts "
            "(needlepoint, woodworking), volunteering/community work, non-medical awards or press, "
            "languages. Include an item ONLY if publicly stated and clearly about THIS physician; "
            "when unsure, omit it. Do NOT include anything sensitive: health, religion, politics, "
            "sexual orientation, race/ethnicity, family members' private details, home address, or "
            "finances.")
        schema.append('"interests": [{"text": <short factual phrase>, "source_url": <url or null>}]')
    parts.append("\nReply with STRICT JSON ONLY (no markdown, no prose): {" + ", ".join(schema) + "}.")
    prompt = "\n".join(parts)

    resp = _post("https://api.anthropic.com/v1/messages",
                 json_body={"model": LLM_MODEL_AUX, "max_tokens": 1000,
                            "messages": [{"role": "user", "content": prompt}],
                            "tools": [{"type": "web_search_20250305", "name": "web_search",
                                       "max_uses": 3}]},
                 headers={"x-api-key": ANTHROPIC_API_KEY,
                          "anthropic-version": "2023-06-01",
                          "content-type": "application/json"})
    if not resp:
        prof.notes.append("Open web search was unavailable this run (rate limit or network).")
        return
    try:
        text = "".join(b.get("text", "") for b in resp.get("content", [])
                       if b.get("type") == "text").strip()
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text).strip()
        mm = re.search(r"\{.*\}", text, re.DOTALL)
        data = json.loads(mm.group(0) if mm else text)
    except Exception:  # noqa: BLE001
        prof.notes.append("Open web search ran but its result could not be parsed.")
        return

    if want_location:
        conf = str(data.get("confidence") or "").strip().lower()
        evidence = [str(u).strip() for u in (data.get("evidence_urls") or []) if str(u).strip()][:6]
        prof.web_confidence = conf
        prof.web_evidence = evidence
        # Identity fallback: if NPPES and CMS gave no name, accept the web-confirmed name (only
        # when the search is reasonably confident and cites a source, so we don't invent a person).
        if not prof.first_name and not prof.last_name and conf in ("high", "medium") and evidence:
            wn = re.sub(r"\s+", " ", str(data.get("provider_name") or "").strip())
            wn = re.sub(r",?\s*(MD|DO|MBBS|NP|PA|DPM|DDS|DMD|PhD)\b.*$", "", wn, flags=re.I).strip()
            toks = [t for t in wn.split(" ") if t]
            if len(toks) >= 2:
                prof.last_name = toks[-1]
                prof.first_name = " ".join(toks[:-1])
                if not prof.credential:
                    prof.credential = str(data.get("provider_credential") or "").strip()
                prof.sources_used.append("Open web search (provider identity)")
                prof.notes.append(f"Provider identified via open web search as {prof.full_name}"
                                  f"{(' ' + prof.credential) if prof.credential else ''} "
                                  f"(NPPES returned no name); confidence {conf}.")
        emp = str(data.get("current_employer") or "").strip()
        if emp and emp.lower() not in ("none", "null", "n/a", "unknown"):
            prof.web_current_employer = emp
        web = str(data.get("website") or "").strip()
        if web.startswith("http") and not prof.website:
            prof.website = web
        offices = []
        for x in (data.get("offices") or []):
            if not isinstance(x, dict):
                continue
            line1 = str(x.get("line1") or "").strip()
            city = str(x.get("city") or "").strip()
            state = str(x.get("state") or "").strip().upper()[:2]
            postal = str(x.get("postal") or "").strip()
            if not line1 or not (city or postal):
                continue
            offices.append({"line1": line1, "line2": str(x.get("line2") or "").strip(),
                            "city": city, "state": state, "postal": postal,
                            "phone": str(x.get("phone") or "").strip()})
        if offices and conf in ("high", "medium") and evidence:
            prof.web_offices = offices[:5]
            prof.sources_used.append("Open web search (current practice)")
            prof.notes.append(
                "Open web search located the current practice"
                + (f" at {prof.web_current_employer}" if prof.web_current_employer else "")
                + f" ({offices[0].get('city','')}, {offices[0].get('state','')}); "
                + f"confidence {conf}. Evidence: " + "; ".join(evidence[:3]))
        elif offices:
            prof.notes.append("Open web search returned a possible current location but with low "
                              "confidence or no cited source; not used to override the address.")

    if want_interests:
        items = []
        for it in (data.get("interests") or []):
            if isinstance(it, dict) and str(it.get("text") or "").strip():
                url = str(it.get("source_url") or "").strip()
                items.append({"text": str(it["text"]).strip(),
                              "source": url if url.startswith("http") else "open web search"})
        if _merge_interests(prof, items):
            prof.sources_used.append("Open web search (personal interests)")


def _merge_interests(prof: ProviderProfile, items, default_source: str = "") -> int:
    """Add interest items ({"text","source"}) to the profile, collapsing near-duplicates: if a
    new item's significant words are a subset of an existing item (or vice versa) they are treated
    as the same interest, keeping the more detailed phrasing (e.g. 'runs marathons' is absorbed by
    'Runs marathons (Boston 2023)')."""
    def toks(s):
        return frozenset(t for t in re.split(r"[^a-z0-9]+", (s or "").lower()) if len(t) > 2)
    added = 0
    for it in items:
        text = (it.get("text") or "").strip()
        nt = toks(text)
        if not text or not nt:
            continue
        src = (it.get("source") or default_source or "").strip()
        handled = False
        for existing in prof.personal_interests:
            et = toks(existing["text"])
            if nt <= et:                       # duplicate or less detailed than what we have
                handled = True
                break
            if et < nt:                        # new item is the more detailed version -> upgrade
                existing["text"] = text
                if src.startswith("http") or not (existing.get("source") or ""):
                    existing["source"] = src or existing.get("source", "")
                handled = True
                break
        if handled:
            continue
        prof.personal_interests.append({"text": text, "source": src})
        added += 1
    return added


def fetch_affiliated_org_locations(prof: ProviderProfile) -> None:
    """For an individual (Type 1) physician, find the organization NPI(s) on which
    they are the Authorized Official and merge in that group's practice locations.

    Individual NPPES records almost never carry secondary practice locations; the
    physician's group / organization (Type 2) NPI usually does. We only merge a
    location when the org's Authorized Official is clearly this physician, so a
    co-located but unrelated practice in the same building is not pulled in.
    """
    if prof.enumeration_type not in ("NPI-1", "Individual", ""):
        return
    if not prof.last_name:
        return

    loc = next((a for a in prof.addresses if a.purpose == "LOCATION"), None) or \
        (prof.addresses[0] if prof.addresses else None)
    state = (loc.state if loc else "") or ""

    have = {_addr_key(a.line1, a.postal) for a in prof.addresses}

    # Two complementary NPPES searches; results are unioned then strictly
    # post-filtered on the Authorized Official name regardless of how they matched.
    searches = []
    s1 = {"version": "2.1", "enumeration_type": "NPI-2",
          "last_name": prof.last_name, "name_purpose": "AO", "limit": 50}
    if prof.first_name:
        s1["first_name"] = prof.first_name
    if state:
        s1["state"] = state
    searches.append(s1)
    if loc and loc.zip5:
        searches.append({"version": "2.1", "enumeration_type": "NPI-2",
                         "address_purpose": "LOCATION", "postal_code": loc.zip5,
                         "state": state, "limit": 50})

    candidates: dict[str, dict] = {}
    for params in searches:
        data = _get(NPPES_URL, params=params)
        for rec in (data or {}).get("results", []) or []:
            num = str(rec.get("number", ""))
            if num and num != prof.npi:
                candidates.setdefault(num, rec)
    if not candidates:
        return

    pf = (prof.first_name or "").strip().lower()
    pl = (prof.last_name or "").strip().lower()
    matched = []
    for rec in candidates.values():
        b = rec.get("basic", {}) or {}
        ao_last = (b.get("authorized_official_last_name", "") or "").strip().lower()
        ao_first = (b.get("authorized_official_first_name", "") or "").strip().lower()
        if not ao_last or ao_last != pl:
            continue
        # If we know both first names, require a reasonable match (handles "Sohan" vs
        # "Sohan R" without grabbing a different same-surname official).
        if pf and ao_first and ao_first != pf and not (
                ao_first.startswith(pf) or pf.startswith(ao_first)):
            continue
        matched.append(rec)
    if not matched:
        return

    added, org_names = 0, []
    for rec in matched:
        b = rec.get("basic", {}) or {}
        org_name = (b.get("organization_name", "") or b.get("name", "")
                    or f"NPI {rec.get('number', '')}")
        org_names.append(f"{org_name} (NPI {rec.get('number', '')})")
        if not prof.affiliated_org_name and not org_name.startswith("NPI "):
            prof.affiliated_org_name = org_name
        # National brand / management company: record the affiliation but do NOT import
        # its nationwide location list as this doctor's offices.
        if any(b in org_name.lower() for b in NATIONAL_BRANDS):
            continue
        rows = [("LOCATION", ad) for ad in (rec.get("addresses", []) or [])
                if ad.get("address_purpose") == "LOCATION"]
        rows += [("PRACTICE LOCATION", ad) for ad in (rec.get("practiceLocations", []) or [])]
        for purpose, ad in rows:
            a = Address(
                purpose=purpose, line1=ad.get("address_1", ""),
                line2=ad.get("address_2", ""), city=ad.get("city", ""),
                state=ad.get("state", ""), postal=ad.get("postal_code", ""),
                phone=ad.get("telephone_number", ""),
                source=f"NPPES (org: {org_name})",
            )
            k = _addr_key(a.line1, a.postal)
            if not a.line1 or k in have:
                continue
            have.add(k)
            prof.addresses.append(a)
            added += 1

    prof.sources_used.append("NPPES affiliated organization NPI(s)")
    prof.notes.append(
        f"Affiliated organization NPI matched ({'; '.join(org_names)}). "
        f"Added {added} practice location(s) registered to the group but not to the "
        "individual NPI. Confirm each is still an active office.")


# ----------------------------------------------------------------------------
# CMS Medicare (by Provider)
# ----------------------------------------------------------------------------
_CMS_UUID_CACHE: dict[str, Any] = {}


def discover_cms_dataset_uuid(override: str = ""):
    if override:
        return override, None
    if "uuid" in _CMS_UUID_CACHE:
        return _CMS_UUID_CACHE["uuid"], _CMS_UUID_CACHE.get("year")
    catalog = _get(CMS_DATA_JSON, timeout=40)
    best = None
    if catalog and isinstance(catalog.get("dataset"), list):
        for ds in catalog["dataset"]:
            title = ds.get("title") or ""
            t = title.lower()
            if ("by provider" not in t or "physician" not in t
                    or "and service" in t or "geography" in t):
                continue
            ym = re.search(r"(20\d{2})", title)
            year = int(ym.group(1)) if ym else 0
            uuid = None
            for dist in ds.get("distribution", []) or []:
                for key in ("accessURL", "downloadURL"):
                    m = re.search(r"/dataset/([0-9a-fA-F-]{36})/", dist.get(key, "") or "")
                    if m:
                        uuid = m.group(1)
                        break
                if uuid:
                    break
            if uuid and (best is None or year > best[0]):
                best = (year, uuid)
    if best:
        _CMS_UUID_CACHE["uuid"], _CMS_UUID_CACHE["year"] = best[1], best[0] or None
        return best[1], (best[0] or None)
    return None, None


def fetch_cms_medicare(npi: str, prof: ProviderProfile, override_uuid: str = "") -> None:
    uuid, year = discover_cms_dataset_uuid(override_uuid)
    if not uuid:
        prof.notes.append(
            "CMS: could not auto-discover the 'by Provider' dataset UUID. Pass "
            "--cms-dataset-uuid <uuid> (from the data.cms.gov dataset page).")
        return
    url = CMS_DATA_API.format(uuid=uuid)
    rows = _get(url, params={"filter[Rndrng_NPI]": npi, "size": 5}, timeout=40)
    if not rows or not isinstance(rows, list):
        prof.notes.append("CMS: no Medicare utilization row for this NPI (may not bill "
                          "Medicare Part B FFS, or year not yet released).")
        return
    row = None
    for r in rows:
        if str(r.get("Rndrng_NPI", "")).strip() == str(npi):
            row = r
            break
    if row is None:
        sample = rows[0] if rows else {}
        npis_seen = [str(r.get("Rndrng_NPI", "")) for r in rows[:5]]
        print(f"    [warn] CMS: filter returned {len(rows)} row(s), none matching NPI {npi}. "
              f"NPIs in response: {npis_seen}. First-row columns: {list(sample.keys())[:12]}",
              file=sys.stderr)
        prof.notes.append("CMS: no exact NPI match in the Medicare 'by Provider' dataset "
                          "(provider may not bill Medicare Part B FFS).")
        return

    prof.sources_used.append("CMS Medicare Physician & Other Practitioners")
    prof.medicare_data_year = year

    # Name fallback (in case NPPES and DAC both came up empty for this valid NPI).
    if not prof.last_name:
        prof.last_name = str(row.get("Rndrng_Prvdr_Last_Org_Name", "") or "").strip()
    if not prof.first_name:
        prof.first_name = str(row.get("Rndrng_Prvdr_First_Name", "") or "").strip()
    if not prof.credential:
        prof.credential = str(row.get("Rndrng_Prvdr_Crdntls", "") or "").strip()
    if not prof.gender:
        prof.gender = str(row.get("Rndrng_Prvdr_Gndr", "") or "").strip()

    def _num(*keys, cast=float):
        for k in keys:
            if k in row and str(row[k]).strip() not in ("", "None"):
                try:
                    return cast(float(row[k]))
                except (ValueError, TypeError):
                    pass
        return None

    prof.medicare_beneficiaries = _num("Tot_Benes", cast=int)
    prof.medicare_total_services = _num("Tot_Srvcs")
    prof.medicare_total_payment = _num("Tot_Mdcr_Pymt_Amt")
    prof.medicare_drug_beneficiaries = _num("Drug_Tot_Benes", cast=int)
    prof.medicare_medical_beneficiaries = _num("Med_Tot_Benes", cast=int)
    prof.medicare_provider_type = str(row.get("Rndrng_Prvdr_Type", "") or row.get("Provider_Type", ""))


# ----------------------------------------------------------------------------
# CMS Doctors & Clinicians (medical school, graduation year, specialties)
# ----------------------------------------------------------------------------
# Typical U.S. ages used to turn a graduation/enumeration year into an age estimate.
_AGE_AT_MED_GRAD = 26       # ~18 starting college + 4 undergrad + 4 medical school
_AGE_AT_ENUMERATION = 30    # NPIs are typically obtained around end of residency
# NPPES opened in 2005; 2005-2007 saw mass back-enrollment, so an enumeration date in
# those years says nothing about when the clinician actually started practicing.
_NPI_INCEPTION_YEARS = {2005, 2006, 2007}


def _row_get(row: dict, *names):
    """Case-insensitive lookup across possible DAC column spellings."""
    lower = {k.lower(): v for k, v in row.items()}
    for n in names:
        if n in row and str(row[n]).strip():
            return str(row[n]).strip()
        v = lower.get(n.lower())
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def fetch_doctors_and_clinicians(npi: str, prof: ProviderProfile) -> None:
    """CMS 'Doctors and Clinicians' National Downloadable File: medical school,
    graduation year, and Medicare primary/secondary specialties. A clinician can
    appear on several rows (one per enrollment/location); the school/grad-year/
    specialty fields are identical across them, so the first matching row suffices."""
    url = DAC_QUERY_URL.format(ds=DAC_DATASET_ID)
    params = {
        "conditions[0][property]": "NPI",
        "conditions[0][value]": npi,
        "conditions[0][operator]": "=",
        "limit": 10,
    }
    data = _get(url, params=params, timeout=40)
    rows = (data or {}).get("results") if isinstance(data, dict) else None
    if not rows:
        prof.notes.append("CMS Doctors & Clinicians: no record for this NPI (clinician may "
                          "not be enrolled in Medicare; medical school / graduation year "
                          "unavailable from this source).")
        return
    row = next((r for r in rows if str(r.get("NPI", r.get("npi", ""))).strip() == str(npi)), rows[0])
    prof.sources_used.append("CMS Doctors & Clinicians National Downloadable File")

    # Name fallback: NPPES occasionally returns no usable record for a valid NPI. DAC carries the
    # provider's name, so fill anything NPPES left blank rather than aborting with a blank report.
    if not prof.last_name:
        prof.last_name = _row_get(row, "Provider Last Name", "lst_nm", "Last Name")
    if not prof.first_name:
        prof.first_name = _row_get(row, "Provider First Name", "frst_nm", "First Name")
    if not prof.middle_name:
        prof.middle_name = _row_get(row, "Provider Middle Name", "mid_nm")
    if not prof.credential:
        prof.credential = _row_get(row, "Provider Credential Text", "Cred", "cred")
    if not prof.gender:
        prof.gender = _row_get(row, "gndr", "Gender")

    grd = _row_get(row, "Grd_yr", "grd_yr")
    if grd:
        try:
            prof.graduation_year = int(float(grd))
        except (ValueError, TypeError):
            pass

    school = _row_get(row, "Med_sch", "med_sch")
    # CMS frequently codes unknown schools as "OTHER"; treat that as no data so the
    # website analysis (if enabled) can supply a real name instead.
    if school and school.upper() not in ("OTHER", "OTHER SCHOOL", "OTHER MEDICAL SCHOOL"):
        prof.medical_school = school
        prof.medical_school_source = "CMS Doctors & Clinicians"

    pri = _row_get(row, "pri_spec", "Pri_spec")
    if pri:
        prof.dac_primary_specialty = pri.title() if pri.isupper() else pri
    secs = []
    allspec = _row_get(row, "sec_spec_all")
    if allspec:
        secs = [s.strip() for s in re.split(r"[|;,]", allspec) if s.strip()]
    else:
        for k in ("sec_spec_1", "sec_spec_2", "sec_spec_3", "sec_spec_4"):
            v = _row_get(row, k)
            if v:
                secs.append(v)
    seen = set()
    for s in secs:
        label = s.title() if s.isupper() else s
        if label.lower() not in seen and label.lower() != prof.dac_primary_specialty.lower():
            seen.add(label.lower())
            prof.dac_secondary_specialties.append(label)

    # Billing org name + group size (for employment/independence assessment).
    fac = _row_get(row, "Facility Name", "facility_name", "org_nm")
    if fac:
        prof.dac_org_name = fac
    grp = _row_get(row, "num_org_mem", "Num_org_mem")
    if grp:
        try:
            prof.group_size = int(float(grp))
        except (ValueError, TypeError):
            pass


def estimate_age(prof: ProviderProfile) -> None:
    """Estimate the clinician's age. Prefer graduation year; fall back to the NPI
    enumeration year, but never use 2005-2007 enumerations (NPI inception/back-enrollment).
    Any value found on the website (via the LLM pass) takes precedence and is set there."""
    if prof.estimated_age is not None:
        return  # already set (e.g., a real age extracted from the website)
    this_year = dt.date.today().year
    if prof.graduation_year and 1950 <= prof.graduation_year <= this_year:
        prof.estimated_age = (this_year - prof.graduation_year) + _AGE_AT_MED_GRAD
        prof.age_basis = (f"estimate: {this_year - prof.graduation_year} yrs since "
                          f"{prof.graduation_year} med-school graduation + ~{_AGE_AT_MED_GRAD} "
                          "at graduation")
        return
    enum_year = None
    if prof.enumeration_date:
        try:
            enum_year = dt.datetime.strptime(prof.enumeration_date, "%Y-%m-%d").year
        except ValueError:
            enum_year = None
    if enum_year and enum_year not in _NPI_INCEPTION_YEARS and 2004 <= enum_year <= this_year:
        prof.estimated_age = (this_year - enum_year) + _AGE_AT_ENUMERATION
        prof.age_basis = (f"rough estimate: {this_year - enum_year} yrs since {enum_year} NPI "
                          f"enumeration + ~{_AGE_AT_ENUMERATION} typical at enumeration "
                          "(no graduation year available)")
    elif enum_year in _NPI_INCEPTION_YEARS:
        prof.age_basis = (f"not estimated: NPI enumerated in {enum_year} (2005-2007 inception "
                          "years are unreliable for dating a career) and no graduation year found")


# ----------------------------------------------------------------------------
# Census ACS area wealth + geocoding
# ----------------------------------------------------------------------------
_CENSUS_CACHE: dict[str, Any] = {}
_CENSUS_VARS = "NAME,B19013_001E,B19301_001E,B25077_001E"


def census_for_zip(zip5: str, state: str = ""):
    if not zip5:
        return None
    fips = _STATE_FIPS.get((state or "").strip().upper(), "")
    cache_key = f"{fips}:{zip5}"
    if cache_key in _CENSUS_CACHE:
        return _CENSUS_CACHE[cache_key]
    last_raw = None
    for year in CENSUS_YEARS_TO_TRY:
        # ZCTAs are only queryable at the national level in 2020+ ACS5; nesting them
        # in a state (&in=state:FIPS) is an unsupported hierarchy and returns HTTP 400.
        attempts = [{"get": _CENSUS_VARS, "for": f"zip code tabulation area:{zip5}"}]
        for params in attempts:
            if CENSUS_API_KEY:
                params["key"] = CENSUS_API_KEY
            raw = _get(CENSUS_BASE.format(year=year), params=params, timeout=30, expect="text")
            if raw:
                last_raw = (year, raw.strip()[:200])
            if not raw or not raw.lstrip().startswith("["):
                continue
            try:
                data = json.loads(raw)
            except Exception:  # noqa: BLE001
                continue
            if isinstance(data, list) and len(data) >= 2:
                rec = dict(zip(data[0], data[1]))

                def _int(v):
                    try:
                        iv = int(float(v))
                        return iv if iv >= 0 else None
                    except (TypeError, ValueError):
                        return None

                out = {"year": year,
                       "median_household_income": _int(rec.get("B19013_001E")),
                       "per_capita_income": _int(rec.get("B19301_001E")),
                       "median_home_value": _int(rec.get("B25077_001E"))}
                _CENSUS_CACHE[cache_key] = out
                return out
    _CENSUS_CACHE[cache_key] = None
    if last_raw:
        print(f"    [warn] Census: no usable ACS row for ZCTA {zip5} "
              f"(last try {last_raw[0]}): {last_raw[1]}", file=sys.stderr)
    else:
        print(f"    [info] Census: no ZCTA-level data for ZIP {zip5} "
              f"(often a PO-box or non-residential ZIP with no tabulation area).",
              file=sys.stderr)
    return None


def enrich_area_wealth(prof: ProviderProfile) -> None:
    if not CENSUS_API_KEY:
        prof.notes.append("Census: area income/home value need a (free) Census API key. "
                          "Get one at https://api.census.gov/data/key_signup.html and set "
                          "it as the CENSUS_API_KEY environment variable.")
        return
    used = False
    for a in prof.addresses:
        if not a.is_office:
            continue  # mailing / PO-box ZIPs have no ZCTA and aren't shown as offices
        info = census_for_zip(a.zip5, a.state)
        if info:
            a.area_median_household_income = info["median_household_income"]
            a.area_per_capita_income = info["per_capita_income"]
            a.area_median_home_value = info["median_home_value"]
            a.census_year = info["year"]
            used = True
    if used:
        prof.sources_used.append("U.S. Census ACS 5-year")


def census_geocode(addr: Address) -> None:
    oneline = addr.one_line()
    if not oneline:
        return
    data = _get(CENSUS_GEOCODER, params={"address": oneline,
                "benchmark": "Public_AR_Current", "format": "json"}, timeout=30)
    try:
        c = data["result"]["addressMatches"][0]["coordinates"]
        addr.lat, addr.lng = float(c["y"]), float(c["x"])
    except (TypeError, KeyError, IndexError, ValueError):
        pass


def google_geocode(addr: Address) -> bool:
    """Precise geocode via Google (the key already covers the Geocoding API)."""
    if not GOOGLE_API_KEY:
        return False
    oneline = addr.one_line()
    if not oneline:
        return False
    data = _get("https://maps.googleapis.com/maps/api/geocode/json",
                params={"address": oneline, "key": GOOGLE_API_KEY}, timeout=25)
    try:
        g = data["results"][0]["geometry"]["location"]
        addr.lat, addr.lng = float(g["lat"]), float(g["lng"])
        return True
    except (TypeError, KeyError, IndexError, ValueError):
        return False


# ----------------------------------------------------------------------------
# Imagery
# ----------------------------------------------------------------------------
def static_map_b64(lat: float, lng: float) -> Optional[str]:
    if GOOGLE_API_KEY:
        gurl = (f"https://maps.googleapis.com/maps/api/staticmap?center={lat},{lng}"
                f"&zoom=16&size=640x320&scale=2&markers=color:red%7C{lat},{lng}"
                f"&key={GOOGLE_API_KEY}")
        b = fetch_image_b64(gurl)
        if b:
            return b
    osm = (f"{OSM_STATICMAP}?center={lat},{lng}&zoom=16&size=640x320"
           f"&markers={lat},{lng},red-pushpin")
    return fetch_image_b64(osm)


def enrich_imagery(prof: ProviderProfile) -> None:
    used = False
    for a in prof.addresses:
        if not a.is_office:
            continue
        if a.lat is None or a.lng is None:
            if not google_geocode(a):
                census_geocode(a)
        if a.lat is None or a.lng is None:
            continue
        a.map_link = f"https://www.openstreetmap.org/?mlat={a.lat}&mlon={a.lng}#map=17/{a.lat}/{a.lng}"
        if not a.map_data_uri:
            mb = static_map_b64(a.lat, a.lng)
            if mb:
                a.map_data_uri = mb
                used = True
        if GOOGLE_API_KEY and not a.street_view_data_uri:
            # Pass the address (not raw lat/lng) so Street View tends to face the
            # premises rather than point off down the street.
            sv_loc = urllib.parse.quote_plus(a.one_line() or f"{a.lat},{a.lng}")
            sv = fetch_image_b64(
                f"https://maps.googleapis.com/maps/api/streetview?size=640x400&fov=80"
                f"&location={sv_loc}&return_error_code=true&key={GOOGLE_API_KEY}")
            if sv:
                a.street_view_data_uri = sv
                used = True
        if GOOGLE_API_KEY and not a.satellite_data_uri:
            # Zoomed aerial: shows the building, parking, and surrounding development
            # so you can judge whether the location fits a concierge practice.
            sat = fetch_image_b64(
                f"https://maps.googleapis.com/maps/api/staticmap?center={a.lat},{a.lng}"
                f"&zoom=17&size=640x400&scale=2&maptype=satellite"
                f"&markers=color:red%7C{a.lat},{a.lng}&key={GOOGLE_API_KEY}")
            if sat:
                a.satellite_data_uri = sat
                used = True
    if used:
        prof.sources_used.append("Geocoding + embedded map/Street View/aerial imagery")


# ----------------------------------------------------------------------------
# Google Places (rating, website, photos) - searches by physician NAME
# ----------------------------------------------------------------------------
def google_places_enrich(prof: ProviderProfile) -> None:
    if not GOOGLE_API_KEY:
        prof.notes.append("Google: set GOOGLE_MAPS_API_KEY to add an aggregate Google "
                          "rating, website, and building photos / Street View.")
        return
    loc = next((a for a in prof.addresses if a.purpose == "LOCATION"), None) or \
        (prof.addresses[0] if prof.addresses else None)
    if not loc:
        return

    field_mask = ("places.id,places.displayName,places.formattedAddress,places.rating,"
                  "places.userRatingCount,places.websiteUri")

    name = prof.full_name
    cred = prof.credential or ""
    spec = prof.primary_taxonomy or prof.dac_primary_specialty or ""
    city = loc.city or ""
    st = loc.state or ""
    # NPPES address fields are frequently years out of date, so we do NOT seed the search with
    # the NPPES city — that just biases the result back toward a stale location. The doctor's
    # CURRENT employer/group (from CMS, fetched just above) is the better signal for where they
    # practice now; we lead with it and let the live listing report the current city.
    current_org = (prof.web_current_employer or prof.dac_org_name or prof.affiliated_org_name
                   or prof.organization_name or "").strip()
    raw_queries = [
        f"{name} {cred} {current_org}",     # name + current employer/group (no stale city)
        f"{name} {spec}",                   # name + specialty, location-agnostic
        f"{name} {cred}",                   # name only
        f"{name} {spec} {city} {st}",       # NPPES city ONLY as a last resort
    ]
    seen, queries = set(), []
    for q in raw_queries:
        q = re.sub(r"\s+", " ", q).strip()
        if q and q.lower() not in seen:
            seen.add(q.lower())
            queries.append(q)

    print(f"    [info] Google Places: trying name-based queries for {name}", file=sys.stderr)
    place = None
    last_resp_empty = True
    for q in queries:
        if not q:
            continue
        resp = _post("https://places.googleapis.com/v1/places:searchText",
                     json_body={"textQuery": q, "maxResultCount": 5},
                     headers={"X-Goog-Api-Key": GOOGLE_API_KEY,
                              "X-Goog-FieldMask": field_mask})
        places = (resp or {}).get("places") or []
        if not places:
            continue
        last_resp_empty = False
        rated = [p for p in places if p.get("rating") is not None]
        place = rated[0] if rated else places[0]
        if rated:
            break

    if not place:
        if last_resp_empty:
            prof.notes.append("Google Places: no listing matched this physician by name "
                              "(the doctor may not have a Google Business listing).")
        return

    prof.sources_used.append("Google Places")
    if not prof.places_name:
        prof.places_name = (place.get("displayName") or {}).get("text", "")
    if place.get("rating") is not None:
        prof.reviews.append(ReviewSource(
            source="Google", rating=place.get("rating"),
            review_count=place.get("userRatingCount"),
            url=f"https://www.google.com/maps/place/?q=place_id:{place.get('id','')}"))
    else:
        prof.notes.append("Google Places matched a listing but it has no rating yet.")
    if place.get("websiteUri") and not prof.website:
        prof.website = place["websiteUri"]
    if place.get("formattedAddress"):
        prof.places_formatted_address = place["formattedAddress"]
    # We deliberately do not copy the listing's coordinates onto the office address,
    # nor pull its user-uploaded photos (often decor/random snapshots, not the
    # building). All imagery is anchored to the registered practice address below
    # so the Street View, aerial, and map agree on the same location.


# ----------------------------------------------------------------------------
# Web scraping: reviews, board certification, recognition
# ----------------------------------------------------------------------------
def _page_text_and_jsonld(html: str):
    jsonld = []
    if _HAVE_BS4:
        try:
            soup = BeautifulSoup(html, "html.parser")
            for s in soup.find_all("script", type="application/ld+json"):
                try:
                    jsonld.append(json.loads(s.string or ""))
                except Exception:  # noqa: BLE001
                    pass
            for tag in soup(["script", "style", "noscript"]):
                tag.decompose()
            text = soup.get_text(" ", strip=True)
        except Exception:  # noqa: BLE001
            text = html
    else:
        text = re.sub(r"<[^>]+>", " ", html)
    return text, jsonld


def _aggregate_rating_from_jsonld(jsonld_objs) -> Optional[ReviewSource]:
    def walk(o):
        if isinstance(o, dict):
            if "aggregateRating" in o and isinstance(o["aggregateRating"], dict):
                ar = o["aggregateRating"]
                try:
                    rating = float(ar.get("ratingValue"))
                except (TypeError, ValueError):
                    rating = None
                cnt = ar.get("reviewCount") or ar.get("ratingCount")
                try:
                    cnt = int(cnt)
                except (TypeError, ValueError):
                    cnt = None
                if rating is not None:
                    return rating, cnt
            for v in o.values():
                res = walk(v)
                if res:
                    return res
        elif isinstance(o, list):
            for v in o:
                res = walk(v)
                if res:
                    return res
        return None

    for obj in jsonld_objs:
        res = walk(obj)
        if res:
            return ReviewSource(source="Website (schema.org)", rating=res[0],
                                review_count=res[1], note="auto-extracted")
    return None


def _ratings_from_text(text: str) -> Optional[ReviewSource]:
    m = re.search(r"(\d(?:\.\d)?)\s*(?:out of|/)\s*5(?:\D{0,40}?(\d{1,4})\s+reviews?)?", text)
    if not m:
        m = re.search(r"(\d(?:\.\d)?)\s*stars?\D{0,40}?(\d{1,4})\s+reviews?", text)
    if m:
        try:
            rating = float(m.group(1))
        except ValueError:
            return None
        cnt = None
        if m.lastindex and m.lastindex >= 2 and m.group(2):
            try:
                cnt = int(m.group(2))
            except ValueError:
                cnt = None
        if 0 <= rating <= 5:
            return ReviewSource(source="Website (text)", rating=rating,
                                review_count=cnt, note="auto-extracted; verify")
    return None


def _find_certs_and_recognition(text: str):
    certs, recog = set(), set()
    for pat in _CERT_PATTERNS:
        for m in re.finditer(pat, text, flags=re.IGNORECASE):
            snippet = re.sub(r"\s+", " ", m.group(0)).strip()
            if 4 < len(snippet) < 120:
                certs.add(snippet.title())
    for pat in _RECOGNITION_PATTERNS:
        for m in re.finditer(pat, text, flags=re.IGNORECASE):
            snippet = re.sub(r"\s+", " ", m.group(0)).strip()
            if 3 < len(snippet) < 80:
                recog.add(snippet.title())
    return sorted(certs)[:8], sorted(recog)[:8]


def discover_website(prof: ProviderProfile) -> None:
    if prof.website:
        return
    # Anchor the manual-lookup hint on the current employer/group rather than the NPPES city,
    # which is often outdated.
    org = (prof.dac_org_name or prof.affiliated_org_name or prof.organization_name or "").strip()
    q = urllib.parse.quote_plus(
        f"{prof.full_name} {prof.primary_taxonomy} {org} official website".strip())
    prof.notes.append(f"Website not auto-found. Lookup: https://duckduckgo.com/?q={q}")


# Link text / href hints for the pages that carry provider rosters and office photos.
_SITE_PAGE_HINTS = ("about", "team", "provider", "providers", "physician", "doctors",
                    "staff", "meet", "our-team", "our-providers", "gallery", "office",
                    "tour", "facility", "location")
_IMG_SKIP = re.compile(r"(logo|icon|sprite|favicon|badge|placeholder|spinner|loader|"
                       r"pixel|spacer|blank|bg-|background|banner-ad)", re.IGNORECASE)


def _collect_site_pages(start_url: str, max_pages: int = 4):
    """Fetch the homepage plus a few internal team/about/provider/gallery pages.
    Returns [(page_url, html)]."""
    pages = []
    home = _get(start_url, expect="text", timeout=20, retries=2)
    if not home:
        return pages
    pages.append((start_url, home))
    if not _HAVE_BS4:
        return pages
    try:
        soup = BeautifulSoup(home, "html.parser")
        base = urllib.parse.urlparse(start_url)
        seen = {start_url.rstrip("/")}
        cand = []
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            txt = (a.get_text(" ", strip=True) or "").lower()
            full = urllib.parse.urljoin(start_url, href)
            pu = urllib.parse.urlparse(full)
            if pu.scheme not in ("http", "https") or pu.netloc != base.netloc:
                continue
            key = full.split("#")[0].rstrip("/")
            if key in seen:
                continue
            hay = (pu.path + " " + txt).lower()
            if any(h in hay for h in _SITE_PAGE_HINTS):
                seen.add(key)
                cand.append(key)
        for url in cand[:max_pages]:
            h = _get(url, expect="text", timeout=15, retries=1, quiet=True)
            if h:
                pages.append((url, h))
    except Exception:  # noqa: BLE001
        pass
    return pages


def _image_candidates(pages, limit: int = 24):
    """Pull plausible content images (not logos/icons) with context, across pages."""
    if not _HAVE_BS4:
        return []
    out, seen = [], set()
    for page_url, html in pages:
        try:
            soup = BeautifulSoup(html, "html.parser")
        except Exception:  # noqa: BLE001
            continue
        for img in soup.find_all("img"):
            src = (img.get("src") or img.get("data-src") or img.get("data-lazy-src")
                   or img.get("data-original") or "")
            if not src or src.startswith("data:"):
                continue
            full = urllib.parse.urljoin(page_url, src.strip())
            if not full.lower().startswith("http") or full.lower().endswith(".svg"):
                continue
            if _IMG_SKIP.search(full):
                continue
            try:
                w = int(re.sub(r"\D", "", str(img.get("width") or "0")) or 0)
                h = int(re.sub(r"\D", "", str(img.get("height") or "0")) or 0)
            except ValueError:
                w = h = 0
            if (w and w < 80) or (h and h < 80):
                continue
            key = full.split("?")[0]
            if key in seen:
                continue
            seen.add(key)
            alt = (img.get("alt") or "").strip()
            near = ""
            fig = img.find_parent("figure")
            if fig:
                cap = fig.find("figcaption")
                if cap:
                    near = cap.get_text(" ", strip=True)[:80]
            if not near:
                prev = img.find_previous(["h1", "h2", "h3", "h4"])
                if prev:
                    near = prev.get_text(" ", strip=True)[:80]
            out.append({"url": full, "alt": alt[:120], "near": near})
            if len(out) >= limit:
                return out
    return out


def _embed_images(urls, cap: int):
    out = []
    for u in urls[:cap]:
        b = fetch_image_b64(u, max_bytes=3_000_000)
        if b:
            out.append((u, b))
    return out


_NAME_STOP = {"dr", "mr", "mrs", "ms", "md", "do", "facp", "faafp", "faap", "mph",
              "mba", "phd", "np", "pa", "pac", "fnp", "aprn", "the", "our", "meet"}


def _name_key(label: str) -> str:
    toks = [t for t in re.findall(r"[a-z]+", (label or "").lower()) if t not in _NAME_STOP]
    return "".join(sorted(toks))


def _dedupe_provider_photos(prov):
    """Drop duplicate people (same normalized name) and duplicate image URLs —
    at most one photo per physician/provider."""
    seen_names, seen_urls, out = set(), set(), []
    for label, url in prov:
        if url in seen_urls:
            continue
        nk = _name_key(label)
        if nk and nk in seen_names:
            continue
        seen_urls.add(url)
        if nk:
            seen_names.add(nk)
        out.append((label, url))
    return out


def analyze_website_with_llm(prof: ProviderProfile, page_text: str, image_candidates=None) -> bool:
    """Read the site text with Claude and extract structured practice info, the midlevel
    roster, and (from a provided candidate list) which images are interior vs provider
    photos. Returns True if it populated fields, False to fall back to keyword matching."""
    image_candidates = image_candidates or []
    if not ANTHROPIC_API_KEY:
        prof.notes.append("Website analysis: set ANTHROPIC_API_KEY to identify the practice "
                          "type, services, midlevel staff, and photos by reading the site "
                          "(instead of crude keyword matching).")
        return False
    snippet = re.sub(r"\s+", " ", page_text)[:8000]
    img_block = ""
    if image_candidates:
        lines = [f'{i}. url={c["url"]} | alt="{c["alt"]}" | near="{c["near"]}"'
                 for i, c in enumerate(image_candidates[:12])]
        img_block = ("\n\nCANDIDATE IMAGES (choose ONLY from these exact URLs; do not invent "
                     "URLs):\n" + "\n".join(lines))
    instructions = (
        "You are analyzing a medical provider's website to help assess it as a "
        "concierge-medicine acquisition target. Using ONLY what the page states or clearly "
        "implies, return STRICT JSON (no markdown, no prose) with these keys: "
        '"practice_type" (short label), "summary" (one sentence), "services" (list), '
        '"insurances" (list; [] if cash-pay/none stated), "accepts_insurance" (true/false/null), '
        '"board_certifications" (list), "recognitions" (list), '
        '"medical_school" (or null), "graduation_year" (4-digit or null), '
        '"residency" (residency training as stated on the page — program/specialty and/or '
        'hospital, e.g. "Internal Medicine, Johns Hopkins Hospital"; null if not stated), '
        '"tenure_at_location" (how long the physician has practiced at THIS office or in the '
        'immediate local area — same city/neighborhood, roughly a 10-minute drive — if the page '
        'states it, e.g. "in practice at this location since 2008" or "serving the community for '
        'over 20 years"; null if not stated or if it clearly refers to a different/distant city), '
        '"age" (only if explicitly stated, else null), '
        '"practice_name" (the practice/group brand, NOT the doctor\'s personal name; or null), '
        '"practice_addresses" (the practice\'s CURRENT office street address(es) exactly as shown '
        'on the site, MOST-LIKELY-PRIMARY first, each as {"line1":..,"line2":.. or null,'
        '"city":..,"state":<2-letter>,"postal":..,"phone":.. or null}; [] if the site shows no '
        "street address. Use only real street addresses for this practice — not a billing/PO-box, "
        "a hospital the doctor merely visits, or another business in the building), "
        '"employer" (named hospital/health system that EMPLOYS the doctor, else null), '
        '"personal_interests" (a list of short PUBLIC, non-clinical personal facts stated on the '
        'page that are useful for rapport — hobbies, sports, music, volunteering, languages, '
        'non-medical interests, e.g. "runs marathons", "plays guitar in a local band"; exclude '
        "anything sensitive such as health, religion, politics, or family details; [] if none), "
        '"midlevels" (a list of the practice\'s mid-level clinicians — count ONLY nurse '
        "practitioners and physician assistants — each as {\"name\":..,\"type\":\"NP\" or "
        "\"PA\"}. Nurse practitioners include credentials NP, APRN, FNP, PMHNP, AGPCNP, "
        "AGACNP, WHNP, PNP, NNP, ENP, DNP (map all of these to \"NP\"). Physician assistants "
        "include PA, PA-C, DMSc, DHSc (map all of these to \"PA\"). Do NOT include physicians, "
        "CNMs, CRNAs, CNSs, RNs, medical assistants, or other staff. [] if the roster is shown "
        "and there are no NPs/PAs), "
        '"provider_roster_visible" (true ONLY if the page clearly shows the full list of the '
        "practice's clinicians so a midlevel count can be trusted; false if you are unsure), "
        '"interior_photo_urls" (URLs FROM THE CANDIDATE LIST showing the office interior — '
        "waiting room, reception, exam rooms, facility; [] if none/none provided), "
        '"provider_photo_urls" (list of {"name":.., "url":..} for headshots of the doctor or '
        "mid-level providers, URLs FROM THE CANDIDATE LIST; [] if none. Include a URL ONLY if "
        "it is a photograph of a real human face or portrait — never scanned documents, "
        "diplomas, certificates, ID cards, logos, maps, or images of text). "
        "Exclude logos, stock photos, and unrelated images. If something is not stated, use [] "
        "or null. Never invent.\n\n"
        f"PAGE TEXT:\n{snippet}{img_block}"
    )
    resp = _post("https://api.anthropic.com/v1/messages",
                 json_body={"model": LLM_MODEL, "max_tokens": 1500,
                            "messages": [{"role": "user", "content": instructions}]},
                 headers={"x-api-key": ANTHROPIC_API_KEY,
                          "anthropic-version": "2023-06-01",
                          "content-type": "application/json"})
    if not resp:
        return False
    try:
        out = "".join(b.get("text", "") for b in resp.get("content", [])
                      if b.get("type") == "text").strip()
        out = re.sub(r"^```(?:json)?\s*|\s*```$", "", out).strip()
        data = json.loads(out)
    except Exception:  # noqa: BLE001
        print("    [warn] LLM website analysis: could not parse JSON response.", file=sys.stderr)
        return False

    def _strlist(v, n):
        return [str(x).strip() for x in v if str(x).strip()][:n] if isinstance(v, list) else []

    prof.practice_type = str(data.get("practice_type") or "").strip()
    prof.practice_summary = str(data.get("summary") or "").strip()
    prof.website_services = _strlist(data.get("services"), 30)
    prof.website_insurances = _strlist(data.get("insurances"), 30)
    prof.board_certifications = _strlist(data.get("board_certifications"), 12)
    prof.recognitions = _strlist(data.get("recognitions"), 12)

    # Website-stated facts take precedence over CMS / estimates.
    ms = str(data.get("medical_school") or "").strip()
    if ms:
        prof.medical_school = ms
        prof.medical_school_source = "provider website"
    res = str(data.get("residency") or "").strip()
    if res and res.lower() not in ("none", "null", "n/a"):
        prof.residency = res
        prof.residency_source = "provider website"
    ten = str(data.get("tenure_at_location") or "").strip()
    if ten and ten.lower() not in ("none", "null", "n/a"):
        prof.tenure_at_location = ten
        prof.tenure_source = "provider website"
    gy = data.get("graduation_year")
    try:
        gy = int(float(gy)) if gy not in (None, "") else None
    except (ValueError, TypeError):
        gy = None
    if gy and 1950 <= gy <= dt.date.today().year:
        prof.graduation_year = gy
    age = data.get("age")
    try:
        age = int(float(age)) if age not in (None, "") else None
    except (ValueError, TypeError):
        age = None
    if age and 25 <= age <= 100:
        prof.estimated_age = age
        prof.age_basis = "stated on provider website"
    pn = str(data.get("practice_name") or "").strip()
    if pn and pn.lower() != prof.full_name.strip().lower():
        prof.practice_name = pn
    offices = []
    for x in (data.get("practice_addresses") or []):
        if not isinstance(x, dict):
            continue
        line1 = str(x.get("line1") or "").strip()
        city = str(x.get("city") or "").strip()
        state = str(x.get("state") or "").strip().upper()[:2]
        postal = str(x.get("postal") or "").strip()
        if not line1 or not (city or postal):
            continue
        offices.append({"line1": line1, "line2": str(x.get("line2") or "").strip(),
                        "city": city, "state": state, "postal": postal,
                        "phone": str(x.get("phone") or "").strip()})
    prof.website_offices = offices[:5]
    ai = data.get("accepts_insurance")
    if isinstance(ai, bool):
        prof.accepts_insurance = ai
    emp = str(data.get("employer") or "").strip()
    if emp and emp.lower() not in ("none", "null", "n/a"):
        prof.employment_status = "likely_employed"
        prof.employment_evidence = f"website states affiliation/employment with {emp}"

    site_interests = [{"text": str(x).strip()} for x in (data.get("personal_interests") or [])
                      if str(x).strip()]
    if site_interests:
        _merge_interests(prof, site_interests, default_source="provider website")

    # Midlevels — count ONLY NPs and PAs; everything else is dropped.
    mids = []
    for x in (data.get("midlevels") or []):
        if isinstance(x, dict):
            t = _normalize_midlevel_type(x.get("type", ""))
            if t:
                mids.append((str(x.get("name", "")).strip() or "(unnamed)", t))
    roster_visible = bool(data.get("provider_roster_visible"))
    prof.midlevels = mids
    if roster_visible:
        prof.midlevel_count = len(mids)
        prof.midlevel_roster_known = True
    else:
        prof.midlevel_count = None
        prof.midlevel_roster_known = False

    # Images — only accept URLs that were in the candidate list.
    if image_candidates:
        allowed = {c["url"] for c in image_candidates}
        interior = [u for u in (data.get("interior_photo_urls") or [])
                    if isinstance(u, str) and u in allowed]
        prov = []
        for x in (data.get("provider_photo_urls") or []):
            if isinstance(x, dict) and x.get("url") in allowed:
                prov.append((str(x.get("name", "")).strip() or "Provider", x["url"]))
        prof.interior_photos = [b for _, b in _embed_images(interior, 4)]
        prof.provider_photos = [(lbl, b) for (lbl, u), (_, b)
                                in _zip_embed(_dedupe_provider_photos(prov), 8)]

    prof.sources_used.append(f"Website analysis ({LLM_MODEL})")
    return True


def _zip_embed(labeled_urls, cap: int):
    """[(label,url)] -> [((label,url),(url,data_uri))] embedding each, dropping failures."""
    out = []
    for label, url in labeled_urls[:cap]:
        b = fetch_image_b64(url, max_bytes=3_000_000)
        if b:
            out.append(((label, url), (url, b)))
    return out


_INTERIOR_HINTS = ("office", "lobby", "reception", "waiting", "exam", "interior",
                   "facility", "suite", "tour", "clinic", "room")
_PROVIDER_HINTS = ("dr ", "dr.", "doctor", " md", "m.d", " do", "d.o", "physician",
                   " np", " pa", "provider", "team", "staff", "headshot", "portrait",
                   "founder")


def _normalize_midlevel_type(s) -> Optional[str]:
    """Map a credential string to 'NP' or 'PA' (or None if it is neither).
    Tolerates suffixes: 'FNP-BC' -> NP, 'PA-C' -> PA, 'AGACNP' -> NP."""
    t = re.sub(r"[^A-Z]", "", str(s).upper())
    if not t:
        return None
    for c in sorted((re.sub(r"[^A-Z]", "", x) for x in NP_CREDENTIALS), key=len, reverse=True):
        if t.startswith(c):
            return "NP"
    for c in sorted((re.sub(r"[^A-Z]", "", x) for x in PA_CREDENTIALS), key=len, reverse=True):
        if t.startswith(c):
            return "PA"
    return None


def _vision_keep(data_uris: list, instruction: str, batch_size: int = 3) -> dict:
    """Ask the vision model a keep/drop question about each image. Returns {index: keep_bool}
    for the indices the model judged; missing indices (and any failure) are simply absent, so
    callers default them to keep. Images are sent in small batches so a single request stays
    well under low input-token-per-minute limits; _post waits out any 429 between batches.
    No-op (empty dict) without an API key."""
    if not data_uris or not ANTHROPIC_API_KEY:
        return {}
    keep: dict[int, bool] = {}
    for base in range(0, len(data_uris), batch_size):
        batch = data_uris[base:base + batch_size]
        content = [{"type": "text", "text": instruction +
                    " Return STRICT JSON (no markdown, no prose): a list of {\"index\": <0-based "
                    "int within THIS batch>, \"keep\": <true|false>}, one entry per image, in "
                    "order."}]
        for i, uri in enumerate(batch):
            m = re.match(r"data:(image/[\w.+-]+);base64,(.*)$", uri or "", re.DOTALL)
            if not m:
                content.append({"type": "text", "text": f"Image {i}: (unverifiable encoding)"})
                continue
            content.append({"type": "text", "text": f"Image {i}:"})
            content.append({"type": "image", "source": {
                "type": "base64", "media_type": m.group(1), "data": m.group(2)}})
        resp = _post("https://api.anthropic.com/v1/messages",
                     json_body={"model": LLM_MODEL_AUX, "max_tokens": 400,
                                "messages": [{"role": "user", "content": content}]},
                     headers={"x-api-key": ANTHROPIC_API_KEY,
                              "anthropic-version": "2023-06-01",
                              "content-type": "application/json"})
        if not resp:
            print("    [warn] photo verification batch failed; keeping those photos unverified.",
                  file=sys.stderr)
            continue
        try:
            out = "".join(b.get("text", "") for b in resp.get("content", [])
                          if b.get("type") == "text").strip()
            out = re.sub(r"^```(?:json)?\s*|\s*```$", "", out).strip()
            verdicts = json.loads(out)
        except Exception:  # noqa: BLE001
            print("    [warn] photo verification: could not parse a batch; keeping those photos.",
                  file=sys.stderr)
            continue
        if isinstance(verdicts, list):
            for v in verdicts:
                if isinstance(v, dict) and isinstance(v.get("index"), int):
                    keep[base + v["index"]] = bool(v.get("keep"))
    return keep


def _verify_provider_photos_are_people(prof: ProviderProfile) -> None:
    """Drop any 'provider photo' that is not actually a photograph of a person (the website
    pass occasionally tags a scanned document, diploma, ID, or logo as a headshot). Default to
    keep on any unjudged/failed image so a flaky verdict never erases a legitimate headshot."""
    photos = list(prof.provider_photos or [])
    if not photos or not ANTHROPIC_API_KEY:
        return
    keep = _vision_keep(
        [u for _, u in photos],
        "Each image is a candidate headshot for a medical provider. For EACH image, decide "
        "whether it is a photograph of one or more real human faces (a headshot or portrait of "
        "a person). Mark keep=false for scanned documents, diplomas, certificates, ID cards, "
        "logos, building or office interiors, maps, charts, or anything that is not a photo of "
        "a person.")
    kept = [p for i, p in enumerate(photos) if keep.get(i, True)]
    dropped = len(photos) - len(kept)
    if dropped:
        prof.provider_photos = kept
        prof.notes.append(f"Dropped {dropped} candidate provider photo(s) that image review "
                          "found were not photographs of a person (e.g. a document or logo).")
        print(f"    [photos] removed {dropped} non-person provider image(s).", file=sys.stderr)


def _verify_interior_photos(prof: ProviderProfile) -> None:
    """Drop any 'interior photo' that does not actually show the inside of the practice, so the
    interior gallery never displays a headshot, exterior, logo, or document. Default to keep on
    any unjudged/failed image."""
    photos = list(prof.interior_photos or [])
    if not photos or not ANTHROPIC_API_KEY:
        return
    keep = _vision_keep(
        photos,
        "Each image is a candidate photo of a medical practice's INTERIOR. For EACH image, "
        "decide whether it actually shows the inside of a clinic or office — waiting room, "
        "reception, exam room, hallway, or similar interior space. Mark keep=false for photos "
        "of people or headshots, exterior building shots, logos, maps, documents, charts, or "
        "stock imagery unrelated to the office.")
    kept = [p for i, p in enumerate(photos) if keep.get(i, True)]
    dropped = len(photos) - len(kept)
    if dropped:
        prof.interior_photos = kept
        prof.notes.append(f"Dropped {dropped} candidate interior photo(s) that image review "
                          "found were not photos of the practice interior.")
        print(f"    [photos] removed {dropped} non-interior image(s).", file=sys.stderr)


def scrape_website(prof: ProviderProfile, do_photos: bool = True) -> None:
    if not prof.website:
        return
    pages = _collect_site_pages(prof.website)
    if not pages:
        return
    prof.sources_used.append("Provider website (scraped)")
    texts, jsonld = [], []
    for _, html in pages:
        t, j = _page_text_and_jsonld(html)
        if t:
            texts.append(t)
        jsonld.extend(j)
    text = "\n".join(texts)
    low = text.lower()
    candidates = _image_candidates(pages) if do_photos else []

    # Preferred path: the model reads the site, extracts structure + midlevels + photos.
    if not analyze_website_with_llm(prof, text, candidates):
        # Keyword fallback (no Anthropic key or call failed).
        prof.website_insurances = sorted({kw for kw in INSURANCE_KEYWORDS if kw in low})
        prof.website_services = sorted({kw for kw in SERVICE_KEYWORDS if kw in low})
        certs, recog = _find_certs_and_recognition(low)
        for c in certs:
            if c not in prof.board_certifications:
                prof.board_certifications.append(c)
        for r in recog:
            if r not in prof.recognitions:
                prof.recognitions.append(r)
        # Midlevels can't be counted reliably without the model -> "not available".
        prof.midlevel_count = None
        prof.midlevel_roster_known = False
        # Heuristic image classification from alt/filename context.
        if do_photos and candidates:
            interior, prov = [], []
            for c in candidates:
                ctx = (c["alt"] + " " + c["near"] + " " + c["url"]).lower()
                if any(h in ctx for h in _INTERIOR_HINTS):
                    interior.append(c["url"])
                elif any(h in ctx for h in _PROVIDER_HINTS):
                    prov.append((c["alt"] or "Provider", c["url"]))
            prof.interior_photos = [b for _, b in _embed_images(interior, 4)]
            prof.provider_photos = [(lbl, b) for (lbl, _), (_, b) in _zip_embed(_dedupe_provider_photos(prov), 8)]

    # Confirm captured headshots are really people, and interior shots really are interiors,
    # before any are displayed.
    _verify_provider_photos_are_people(prof)
    _verify_interior_photos(prof)

    rev = _aggregate_rating_from_jsonld(jsonld) or _ratings_from_text(low)
    if rev:
        rev.url = prof.website
        prof.reviews.append(rev)


# ----------------------------------------------------------------------------
# Physician directory profiles (Healthgrades / U.S. News / Vitals / WebMD / Zocdoc)
# ----------------------------------------------------------------------------
def _cse_search_urls(query: str, domain: str, num: int = 8) -> list[str]:
    """Google Custom Search JSON API, restricted to one site. Reuses GOOGLE_API_KEY."""
    if not (GOOGLE_API_KEY and GOOGLE_CSE_ID):
        return []
    data = _get("https://www.googleapis.com/customsearch/v1",
                params={"key": GOOGLE_API_KEY, "cx": GOOGLE_CSE_ID, "q": query,
                        "siteSearch": domain, "siteSearchFilter": "i", "num": min(num, 10)})
    return [it.get("link", "") for it in (data or {}).get("items", []) if it.get("link")]


def _ddg_search_urls(query: str, domain: str) -> list[str]:
    """DuckDuckGo HTML endpoint (no API key). Returns result URLs on the target domain."""
    html = _get("https://html.duckduckgo.com/html/",
                params={"q": f"{query} site:{domain}"}, expect="text",
                headers={"Referer": "https://duckduckgo.com/"})
    if not html:
        return []
    urls = []
    if _HAVE_BS4:
        soup = BeautifulSoup(html, "html.parser")
        anchors = [a.get("href", "") for a in soup.find_all("a")]
    else:
        anchors = re.findall(r'href="([^"]+)"', html)
    for href in anchors:
        if not href:
            continue
        # DDG wraps external links as /l/?uddg=<encoded real url>
        m = re.search(r"[?&]uddg=([^&]+)", href)
        real = urllib.parse.unquote(m.group(1)) if m else href
        if real.startswith("//"):
            real = "https:" + real
        if domain in real and real.startswith("http"):
            urls.append(real)
    return urls


def _pick_profile_url(urls: list[str], domain: str, last_name: str, path_hints) -> str:
    last = (last_name or "").lower()
    for u in urls:
        try:
            parsed = urllib.parse.urlparse(u)
        except ValueError:
            continue
        if not parsed.netloc.endswith(domain):
            continue
        path = parsed.path.lower()
        # Prefer URLs that look like an individual profile rather than a search/listing page.
        if (last and last in path) or any(h in path for h in path_hints):
            return u
    # Fall back to the first same-domain URL that isn't an obvious search page.
    for u in urls:
        if domain in u and not re.search(r"/search|/find|\?q=", u, re.IGNORECASE):
            return u
    return ""


def fetch_aggregator_reviews(prof: ProviderProfile) -> None:
    """Find the clinician on physician directories, capture each profile link, and
    extract a star rating / review count where the page exposes one (schema.org
    aggregateRating or visible text). Ratings also feed the blended average."""
    if not prof.first_name and not prof.last_name:
        return
    loc = next((a for a in prof.addresses if a.is_office), None) or \
        (prof.addresses[0] if prof.addresses else None)
    city = (loc.city if loc else "") or ""
    st = (loc.state if loc else "") or ""
    query = " ".join(filter(None, [prof.full_name, prof.credential, city, st]))

    found_any = False
    for label, domain, path_hints in AGGREGATOR_SITES:
        urls = _cse_search_urls(query, domain) or _ddg_search_urls(query, domain)
        url = _pick_profile_url(urls, domain, prof.last_name, path_hints)
        if not url:
            continue
        prof.external_profiles.append(ReviewSource(source=label, url=url))
        found_any = True
        # Directory pages are heavy and bot-protected; fetch best-effort (one quiet try)
        # so a slow site degrades to "link captured, rating skipped" without noise.
        html = _get(url, expect="text", timeout=25, retries=1, quiet=True)
        if not html:
            prof.external_profiles[-1].note = "profile found; rating not retrieved (site slow or blocked)"
            continue
        text, jsonld = _page_text_and_jsonld(html)
        rev = _aggregate_rating_from_jsonld(jsonld)
        if rev is None and text:
            rev = _ratings_from_text(text.lower())
        if rev and rev.rating is not None and 0 <= rev.rating <= 5:
            prof.external_profiles[-1].rating = rev.rating
            prof.external_profiles[-1].review_count = rev.review_count
            prof.reviews.append(ReviewSource(source=label, rating=rev.rating,
                                             review_count=rev.review_count, url=url))
    if found_any:
        prof.sources_used.append("Physician directory profiles "
                                 "(Healthgrades / U.S. News / Vitals / WebMD / Zocdoc)")
    else:
        prof.notes.append("Physician directories: no profile auto-located. These sites block "
                          "automated lookups aggressively; setting GOOGLE_CSE_ID improves "
                          "discovery, or search the doctor's name on Healthgrades/U.S. News.")


# ----------------------------------------------------------------------------
# Mosaic consumer-segment enrichment
# ----------------------------------------------------------------------------
_MOSAIC_CACHE: dict = {}  # populated once per process: {"zips":..., "rank":..., "maxlh":...}
_MOSAIC_HIGH_SEGMENTS = {"Proactive", "Healthcare Focused"}


def _cell_str(v) -> str:
    """Coerce any openpyxl cell value to a clean, hashable string. Handles rich-text
    and formula objects that otherwise can't be used as dict keys."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    try:
        from openpyxl.cell.rich_text import CellRichText
        if isinstance(v, CellRichText):
            return "".join(str(t) for t in v).strip()
    except Exception:  # noqa: BLE001
        pass
    return str(v).strip()


def _load_mosaic() -> dict:
    """Load the by-ZIP Mosaic workbook once and cache it. Returns {} if unavailable."""
    if _MOSAIC_CACHE:
        return _MOSAIC_CACHE
    if not _HAVE_OPENPYXL or not Path(MOSAIC_XLSX).exists():
        _MOSAIC_CACHE["zips"] = None
        return _MOSAIC_CACHE
    try:
        wb = openpyxl.load_workbook(MOSAIC_XLSX, read_only=True, data_only=True)
        ws = wb["ZIP_Mosaic_Pop_Added"]
        rows = ws.iter_rows(values_only=True)
        header = [_cell_str(h) for h in next(rows)]   # coerce -> all hashable strings
        meta = {"ZipCode", "Top10Count", "City", "State", "Population",
                "HH_Count", "TT_HH_P", "Avg_HH_Size"}
        mcols = [(i, h) for i, h in enumerate(header) if h and h not in meta]
        zi = header.index("ZipCode")
        ci = header.index("City") if "City" in header else None
        si = header.index("State") if "State" in header else None
        zips = {}
        for r in rows:
            if not r or r[zi] is None:
                continue
            z = _cell_str(r[zi]).split(".")[0].zfill(5)
            counts = {}
            for i, code in mcols:
                v = r[i]
                if v in (None, ""):
                    continue
                try:
                    n = int(float(v))
                except (ValueError, TypeError):
                    continue
                if n > 0:
                    counts[code] = n
            zips[z] = {"counts": counts,
                       "city": _cell_str(r[ci]) if ci is not None else "",
                       "state": _cell_str(r[si]) if si is not None else ""}
        rk = wb["ZIP_Mosaic_Pop_Ranking"]
        rrows = rk.iter_rows(values_only=True)
        rh = [_cell_str(h) for h in next(rrows)]
        mi, ni, li, gi = (rh.index("Mosaic"), rh.index("Name"), rh.index("LH"), rh.index("Segment"))
        rank = {}
        for r in rrows:
            if not r or r[mi] is None:
                continue
            try:
                lh = float(r[li]) if r[li] is not None else 0.0
            except (ValueError, TypeError):
                lh = 0.0
            rank[_cell_str(r[mi])] = {"name": _cell_str(r[ni]), "lh": lh,
                                      "segment": _cell_str(r[gi])}
        wb.close()
        _MOSAIC_CACHE.update(zips=zips, rank=rank,
                             maxlh=max((v["lh"] for v in rank.values()), default=0.21))
    except Exception as e:  # noqa: BLE001
        print(f"    [warn] Mosaic workbook could not be read: {e}", file=sys.stderr)
        _MOSAIC_CACHE["zips"] = None
    return _MOSAIC_CACHE


def _lh_indicator(lh, maxlh: float = 0.21):
    """Convert a proprietary LH strength value into a 1-5 dot indicator + word, so the
    raw Experian number is never displayed."""
    mx = (_MOSAIC_CACHE.get("maxlh") or maxlh) or 0.21
    frac = max(0.0, min(1.0, (lh or 0) / mx))
    level = 0 if not lh else max(1, min(5, round(frac * 5)))
    words = {0: "—", 1: "Very low", 2: "Low", 3: "Moderate", 4: "High", 5: "Very high"}
    return "●" * level + "○" * (5 - level), words[level]


def mosaic_for_zip(zip5: str):
    """Return (top5, score_0_100, high_value_share_pct, classified_pop) for a ZIP, or None."""
    m = _load_mosaic()
    if not m.get("zips") or zip5 not in m["zips"]:
        return None
    rank, maxlh = m["rank"], m["maxlh"] or 0.21
    counts = {c: n for c, n in m["zips"][zip5]["counts"].items()
              if c != "U00" and c in rank}
    tot = sum(counts.values())
    if tot == 0:
        return None
    wlh = sum(n * rank[c]["lh"] for c, n in counts.items()) / tot
    score = round(wlh / maxlh * 100, 1) if maxlh else 0.0
    hv = round(sum(n for c, n in counts.items()
                   if rank[c]["segment"] in _MOSAIC_HIGH_SEGMENTS) / tot * 100, 1)
    top = sorted(counts.items(), key=lambda kv: kv[1], reverse=True)[:5]
    top5 = [(c, rank[c]["name"], rank[c]["segment"], n,
             round(n / tot * 100, 1), rank[c]["lh"]) for c, n in top]
    return top5, score, hv, tot


def enrich_mosaic(prof: ProviderProfile) -> None:
    m = _load_mosaic()
    if not m.get("zips"):
        if not _HAVE_OPENPYXL:
            prof.notes.append("Mosaic: install openpyxl to enable Mosaic consumer scoring.")
        elif not Path(MOSAIC_XLSX).exists():
            prof.notes.append(f"Mosaic: workbook not found at {MOSAIC_XLSX} "
                              "(place ZIP_Mosaic_Pop_Added.xlsx next to main.py or set MOSAIC_XLSX).")
        return
    used = False
    for a in prof.addresses:
        if not a.is_office or not a.zip5:
            continue
        res = mosaic_for_zip(a.zip5)
        if not res:
            continue
        a.mosaic_top5, a.mosaic_score, a.mosaic_high_value_share, a.mosaic_classified_pop = res
        used = True
    if used:
        prof.sources_used.append("Experian Mosaic consumer segmentation (by ZIP)")


# ----------------------------------------------------------------------------
# Market context: concierge competition & affiliate proximity
# ----------------------------------------------------------------------------
def _haversine_mi(lat1, lng1, lat2, lng2) -> float:
    r = 3958.7613  # Earth radius, miles
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lng2 - lng1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(min(1.0, math.sqrt(a)))


def enrich_competition(prof: ProviderProfile) -> None:
    """Count nearby concierge competitors via Google Places (DPC excluded)."""
    if not GOOGLE_API_KEY:
        return
    loc = _primary_office(prof)
    if not loc:
        return
    if loc.lat is None or loc.lng is None:
        if not google_geocode(loc):
            census_geocode(loc)
    if loc.lat is None or loc.lng is None:
        return
    radius_m = min(50000, int(COMPETITION_RADIUS_MI * 1609.34))
    field_mask = "places.id,places.displayName,places.location,places.formattedAddress"
    self_name = prof.full_name.strip().lower()
    found = {}
    for q in COMPETITION_QUERIES:
        resp = _post("https://places.googleapis.com/v1/places:searchText",
                     json_body={"textQuery": q, "maxResultCount": 20,
                                "locationBias": {"circle": {
                                    "center": {"latitude": loc.lat, "longitude": loc.lng},
                                    "radius": radius_m}}},
                     headers={"X-Goog-Api-Key": GOOGLE_API_KEY,
                              "X-Goog-FieldMask": field_mask})
        for pl in (resp or {}).get("places", []) or []:
            pid = pl.get("id")
            name = (pl.get("displayName") or {}).get("text", "")
            low = name.lower()
            loc2 = pl.get("location") or {}
            la, ln = loc2.get("latitude"), loc2.get("longitude")
            if not pid or la is None or ln is None or not name:
                continue
            if not any(k in low for k in COMPETITION_KEYWORDS):
                continue
            if any(x in low for x in COMPETITION_EXCLUDE):
                continue
            if self_name and (self_name in low or low in self_name):
                continue
            dist = _haversine_mi(loc.lat, loc.lng, la, ln)
            if dist <= COMPETITION_RADIUS_MI and pid not in found:
                found[pid] = (name, round(dist, 1))
    comp = sorted(found.values(), key=lambda x: x[1])
    prof.competitor_count = len(comp)
    prof.competitors = comp[:8]
    prof.sources_used.append("Google Places (concierge competition scan)")


_AFFIL_CACHE: list = []


def _load_affiliates() -> list:
    """Load affiliate locations CSV once: returns [(name, lat, lng)]. [] if unavailable."""
    if _AFFIL_CACHE:
        return _AFFIL_CACHE[0]
    out = []
    if Path(AFFILIATES_CSV).exists():
        import csv
        try:
            with open(AFFILIATES_CSV, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                fields = reader.fieldnames or []
                low = {c: c.lower().strip() for c in fields}

                def pick(*subs, exact=()):
                    for c in fields:                       # exact match first
                        if low[c] in exact:
                            return c
                    for c in fields:                       # then substring
                        if any(s in low[c] for s in subs):
                            return c
                    return None

                latc = pick("latitude", "lat", exact=("lat", "latitude", "y"))
                lngc = pick("longitude", "long", "lng", "lon",
                            exact=("lng", "lon", "long", "longitude", "x"))
                ncol = pick("name", "practice", "affiliate", "organization", "physician",
                            "doctor", "provider")
                if latc and lngc:
                    for row in reader:
                        try:
                            la, ln = float(str(row[latc]).strip()), float(str(row[lngc]).strip())
                        except (ValueError, TypeError, KeyError):
                            continue
                        if not (-90 <= la <= 90 and -180 <= ln <= 180):
                            continue
                        out.append((str(row.get(ncol, "") if ncol else "").strip() or "Affiliate", la, ln))
        except Exception as e:  # noqa: BLE001
            print(f"    [warn] Affiliates CSV could not be read: {e}", file=sys.stderr)
    _AFFIL_CACHE.append(out)
    return out


def enrich_affiliate_proximity(prof: ProviderProfile) -> None:
    affs = _load_affiliates()
    if not affs:
        if not Path(AFFILIATES_CSV).exists():
            prof.notes.append(f"Affiliate proximity: no affiliate file at {AFFILIATES_CSV} "
                              "(provide a CSV with name + latitude/longitude columns).")
        return
    loc = _primary_office(prof)
    if not loc:
        return
    if loc.lat is None or loc.lng is None:
        if not google_geocode(loc):
            census_geocode(loc)
    if loc.lat is None or loc.lng is None:
        return
    dists = sorted(((name, _haversine_mi(loc.lat, loc.lng, la, ln)) for name, la, ln in affs),
                   key=lambda x: x[1])
    if dists:
        prof.nearest_affiliate_name = dists[0][0]
        prof.nearest_affiliate_mi = round(dists[0][1], 1)
        prof.affiliates_nearby = sum(1 for _, d in dists if d <= AFFILIATE_NEARBY_MI)
        prof.sources_used.append("Affiliate network proximity")


# ----------------------------------------------------------------------------
# Concierge Fit Score
# ----------------------------------------------------------------------------
def _primary_office(prof: ProviderProfile):
    return next((a for a in prof.addresses if a.is_office and a.purpose == "LOCATION"), None) or \
        next((a for a in prof.addresses if a.is_office), None) or \
        (prof.addresses[0] if prof.addresses else None)


def _specialty_tier(prof: ProviderProfile) -> str:
    hay = " ".join([prof.primary_taxonomy or "", prof.dac_primary_specialty or ""]).lower()
    if not hay.strip():
        return "unknown"
    if any(k in hay for k in SPECIALTY_HIGH):
        return "high"
    if any(k in hay for k in SPECIALTY_OK):
        return "acceptable"
    return "none"


def assess_employment(prof: ProviderProfile) -> None:
    """Set employment_status. A billing-org name matching a hospital/health-system
    pattern (or a website employment claim already captured) => likely_employed."""
    if prof.employment_status == "likely_employed":
        return  # already set from the website
    candidates = [prof.dac_org_name, prof.organization_name, prof.affiliated_org_name]
    for nm in candidates:
        low = (nm or "").lower()
        if not low:
            continue
        hit = next((p for p in EMPLOYER_PATTERNS if p in low), None)
        if hit:
            prof.employment_status = "likely_employed"
            prof.employment_evidence = f"billing/affiliation org '{nm}' matches '{hit}'"
            return
    if (prof.sole_proprietor or "").upper() == "YES":
        prof.employment_status = "independent"
        prof.employment_evidence = "NPPES sole proprietor = YES"
    else:
        prof.employment_status = prof.employment_status or "independent"


def _score_affluence(prof) -> Optional[float]:
    loc = _primary_office(prof)
    inc = loc.area_median_household_income if loc else None
    inc_score = None
    if isinstance(inc, int):
        for thr, sc in [(120000, 100), (90000, 80), (70000, 60),
                        (50000, 40), (35000, 25), (0, 10)]:
            if inc >= thr:
                inc_score = sc
                break
    mos = loc.mosaic_score if loc else None
    parts = [s for s in (inc_score, mos) if s is not None]
    return round(sum(parts) / len(parts), 1) if parts else None


def _score_medicare_volume(prof) -> Optional[float]:
    b = prof.medicare_beneficiaries
    if b is None:
        return None
    # 400+ beneficiaries is a great panel for conversion.
    for thr, sc in [(400, 100), (300, 85), (200, 70), (100, 50), (50, 30), (0, 15)]:
        if b >= thr:
            return sc
    return 15


def _score_career_stage(prof) -> Optional[float]:
    """Centered on the average affiliate age (~61). Plateau = 100; younger doctors
    score progressively lower; 75+ tapers slightly."""
    age = prof.estimated_age
    if age is None:
        return None
    if CAREER_PLATEAU_LOW <= age <= CAREER_PLATEAU_HIGH:
        prof.career_stage = "prime"
        return 100.0
    if age < CAREER_PLATEAU_LOW:
        prof.career_stage = "younger than typical affiliate"
        return max(20.0, round(100 - (CAREER_PLATEAU_LOW - age) * CAREER_YOUNG_SLOPE, 1))
    prof.career_stage = "75+, slightly past typical"
    return max(55.0, round(100 - (age - CAREER_PLATEAU_HIGH) * CAREER_OLD_SLOPE, 1))


def _score_specialty(tier: str) -> Optional[float]:
    return {"high": 100.0, "acceptable": 60.0, "none": 0.0}.get(tier, None)


def compute_fit_score(prof: ProviderProfile) -> None:
    """Blend the weighted components into 0-100 and assign a triage lane.
    Hard disqualifiers force DEAD/0 regardless of the weighted score."""
    prof.panel_estimate = (round(prof.medicare_beneficiaries / 0.30)
                           if prof.medicare_beneficiaries else None)
    tier = _specialty_tier(prof)
    prof.specialty_fit_tier = tier
    assess_employment(prof)

    # Hard disqualifiers (force DEAD): non-target specialty and deactivated NPI.
    dq = []
    if prof.npi_deactivated:
        dq.append("NPI is deactivated/retired")
    if tier == "none" and SPECIALTY_OTHER_IS_DEAD:
        dq.append(f"Primary specialty not a concierge target ({prof.primary_taxonomy or prof.dac_primary_specialty or 'unknown'})")

    # Independence/ownership is now expressed as flags, not a score.
    flags = []
    if prof.employment_status == "likely_employed":
        flags.append(("red", f"Employed by hospital/health system — {prof.employment_evidence}"))
    if prof.group_size is not None and prof.group_size >= 6:
        flags.append(("yellow", f"Group of {prof.group_size} providers — breaking away with the "
                                "panel is harder (not scored)"))
    prof.flags = flags

    prac = (prof.practice_type or "").lower()
    if any(k in prac for k in ("concierge", "membership")):
        prof.acquisition_target = True
        prof.acquisition_reason = ("already an existing concierge/membership practice — prime "
                                   "acquisition target (switch from competitor or independent)")

    subscores = {
        "affluence": _score_affluence(prof),
        "medicare_volume": _score_medicare_volume(prof),
        "career_stage": _score_career_stage(prof),
        "specialty": _score_specialty(tier),
    }
    labels = {"affluence": "Patient affluence & consumer fit",
              "medicare_volume": "Medicare volume",
              "career_stage": "Career stage",
              "specialty": "Specialty fit"}

    # Renormalize weights across components that actually have data, so a missing
    # signal doesn't silently drag the score to zero.
    avail = {k: v for k, v in subscores.items() if v is not None}
    wsum = sum(FIT_WEIGHTS[k] for k in avail) or 1
    total = sum(v * FIT_WEIGHTS[k] for k, v in avail.items()) / wsum
    prof.fit_components = [
        (labels[k], FIT_WEIGHTS[k], (round(subscores[k], 1) if subscores[k] is not None else None),
         (round(subscores[k] * FIT_WEIGHTS[k] / wsum, 1) if subscores[k] is not None else None))
        for k in FIT_WEIGHTS
    ]
    prof.disqualifiers = dq
    if dq:
        prof.fit_score = 0
        prof.fit_lane = "DEAD"
        return
    prof.fit_score = int(round(total))
    prof.fit_lane = next((lane for lane, cut in FIT_LANES if prof.fit_score >= cut), "DEAD")


# ----------------------------------------------------------------------------
# Orchestration
# ----------------------------------------------------------------------------
def filter_offices_to_locality(prof: ProviderProfile) -> None:
    """Keep only office locations at/near the primary practice address. Far-flung
    addresses (historical practices or a national network's other sites) are set
    aside so the report reflects where the doctor actually practices now."""
    offices = [a for a in prof.addresses if a.is_office]
    mailing = [a for a in prof.addresses if not a.is_office]
    if len(offices) <= 1:
        return
    primary = _primary_office(prof)
    if primary is None:
        return
    pcity = (primary.city or "").strip().lower()
    pstate = (primary.state or "").strip().upper()
    pzip3 = (primary.zip5 or "")[:3]
    kept, dropped = [], []
    for a in offices:
        if a is primary:
            kept.append(a)
            continue
        near = False
        if (primary.lat is not None and primary.lng is not None
                and a.lat is not None and a.lng is not None):
            near = _haversine_mi(primary.lat, primary.lng, a.lat, a.lng) <= OFFICE_LOCALITY_MI
        else:
            near = ((a.city or "").strip().lower() == pcity and (a.state or "").strip().upper() == pstate) \
                or (bool(pzip3) and (a.zip5 or "")[:3] == pzip3)
        (kept if near else dropped).append(a)
    if dropped:
        prof.addresses = kept + mailing
        prof.notes.append(
            f"Filtered to {len(kept)} office location(s) within ~{OFFICE_LOCALITY_MI} mi of the "
            f"primary practice address; set aside {len(dropped)} other location(s) (likely "
            "historical or affiliated-network addresses). Verify the current office if needed.")


def _npi_checksum_valid(npi: str) -> bool:
    """Validate a 10-digit NPI with the ISO/IEC 7812 Luhn check (prefix 80840). Catches typos
    and fabricated numbers so we can give a precise message instead of a blank report."""
    s = re.sub(r"\D", "", str(npi or ""))
    if len(s) != 10:
        return False
    base = "80840" + s[:9]
    total = 0
    for i, ch in enumerate(reversed(base)):
        d = int(ch)
        if i % 2 == 0:
            d *= 2
            if d > 9:
                d -= 9
        total += d
    return (10 - (total % 10)) % 10 == int(s[9])


def research_npi(npi: str, *, cms_uuid: str = "", do_google: bool = True,
                 do_scrape: bool = True, do_aggregators: bool = True,
                 do_photos: bool = True, do_web: bool = True,
                 do_interests: bool = True) -> ProviderProfile:
    npi = re.sub(r"\D", "", str(npi))
    prof = ProviderProfile(npi=npi, retrieved_at=dt.datetime.now().isoformat(timespec="seconds"))
    if not _npi_checksum_valid(npi):
        prof.notes.append("Not a valid NPI number: it must be 10 digits and pass the NPI "
                          "checksum. Double-check the number." if len(npi) == 10 else
                          "Not a valid NPI: must be exactly 10 digits.")
        return prof

    print(f"[{npi}] NPPES ...")
    fetch_nppes(npi, prof)
    if not prof.full_name:
        # NPPES sometimes returns nothing for a VALID NPI (deactivated, withheld, or a transient
        # API gap). NPPES is unreliable, so we do NOT abort — CMS and the open web search below
        # usually still identify the provider and build a useful report.
        prof.notes.append("No active NPPES record was returned for this valid NPI (it may be "
                          "deactivated/withheld, or NPPES was briefly unavailable). Continuing "
                          "with CMS and open-web sources to identify the provider.")
    print(f"[{npi}] Affiliated org locations ...")
    fetch_affiliated_org_locations(prof)
    dedupe_addresses(prof)
    print(f"[{npi}] CMS Medicare ...")
    fetch_cms_medicare(npi, prof, override_uuid=cms_uuid)
    print(f"[{npi}] CMS Doctors & Clinicians ...")
    fetch_doctors_and_clinicians(npi, prof)

    # Resolve the CURRENT practice address before any location-based enrichment, since NPPES
    # address data is often years out of date. Order of authority: an open web search for where
    # the doctor practices now, then the practice website, then a Google listing.
    if do_web or do_interests:
        print(f"[{npi}] Open web search (current practice"
              + (" + interests" if do_interests else "") + ") ...")
        find_current_info_via_web(prof, want_location=do_web, want_interests=do_interests)
    if do_google:
        print(f"[{npi}] Google Places ...")
        google_places_enrich(prof)
    discover_website(prof)
    if do_scrape:
        print(f"[{npi}] Website scrape ...")
        scrape_website(prof, do_photos=do_photos)
    apply_current_address(prof)
    dedupe_addresses(prof)

    # Now enrich the resolved current office(s): area wealth, geocoding/imagery, locality.
    print(f"[{npi}] Census area wealth ...")
    enrich_area_wealth(prof)
    print(f"[{npi}] Geocoding + imagery ...")
    enrich_imagery(prof)
    collapse_colocated_offices(prof)
    filter_offices_to_locality(prof)
    _CURRENT_SRC_HINTS = ("website", "open web search", "google business")
    if not any(any(h in (a.source or "").lower() for h in _CURRENT_SRC_HINTS)
               for a in prof.addresses if a.is_office):
        prof.notes.append("Office address could not be confirmed against a live current source "
                          "(web search, practice website, or Google listing); the address shown "
                          "is from NPPES and may lag the practice's current location — verify "
                          "before outreach.")
    if do_aggregators:
        print(f"[{npi}] Physician directories (Healthgrades/U.S. News/...) ...")
        fetch_aggregator_reviews(prof)
    estimate_age(prof)
    print(f"[{npi}] Mosaic consumer segments ...")
    enrich_mosaic(prof)
    print(f"[{npi}] Market context (competition + affiliates) ...")
    enrich_competition(prof)
    enrich_affiliate_proximity(prof)
    compute_fit_score(prof)
    if not prof.reviews:
        prof.notes.append("No aggregate rating could be auto-extracted. If the doctor has a "
                          "Google listing, confirm Places API (New) is enabled and the key is set.")
    return prof


def viability_signals(prof: ProviderProfile) -> dict:
    loc = _primary_office(prof)
    return {
        "office_count": len([a for a in prof.addresses if a.is_office]),
        "area_median_household_income": loc.area_median_household_income if loc else None,
        "area_median_home_value": loc.area_median_home_value if loc else None,
    }


# ----------------------------------------------------------------------------
# Output
# ----------------------------------------------------------------------------
def _profile_for_json(prof: ProviderProfile) -> dict:
    d = asdict(prof)
    for a in d.get("addresses", []):
        n = len(a.get("photo_data_uris") or [])
        a["photo_data_uris"] = f"<{n} embedded photo(s)>" if n else []
        if a.get("street_view_data_uri"):
            a["street_view_data_uri"] = "<embedded>"
        if a.get("satellite_data_uri"):
            a["satellite_data_uri"] = "<embedded>"
        if a.get("map_data_uri"):
            a["map_data_uri"] = "<embedded>"
    return d


def write_json(prof: ProviderProfile, outdir: str) -> str:
    path = os.path.join(outdir, f"{prof.npi}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(_profile_for_json(prof), f, indent=2, ensure_ascii=False)
    return path


def _esc(s: Any) -> str:
    return (str(s) if s is not None else "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def write_html(prof: ProviderProfile, outdir: str) -> str:
    css = """
    body{font-family:Segoe UI,Arial,sans-serif;color:#1b1b1b;margin:0;background:#f4f6fa}
    .wrap{max-width:1000px;margin:0 auto;padding:24px}
    .card{background:#fff;border:1px solid #e1e5ee;border-radius:10px;padding:24px;margin:0 0 22px;box-shadow:0 1px 3px rgba(0,0,0,.05)}
    h1{font-size:21px;margin:0 0 4px} h2{font-size:15px;color:#1F3864;border-bottom:2px solid #1F3864;padding-bottom:4px;margin:24px 0 10px}
    h3{font-size:13px;margin:14px 0 6px;color:#33415c}
    .sub{color:#5a6577;font-size:13px;margin:0 0 14px}
    table{border-collapse:collapse;width:100%;font-size:13px;margin:6px 0}
    td,th{border:1px solid #e1e5ee;padding:6px 9px;text-align:left;vertical-align:top}
    th{background:#f0f3fa;color:#33415c}
    th.key{width:230px}
    .pill{display:inline-block;background:#eaf0fb;color:#1F3864;border-radius:12px;padding:2px 10px;margin:2px 4px 2px 0;font-size:12px}
    .office{border:1px solid #e8ebf2;border-radius:8px;padding:12px;margin:10px 0;background:#fafbfe}
    .imgs{display:flex;flex-wrap:wrap;gap:12px;margin-top:8px}
    .imgs figure{margin:0} .imgs img{height:300px;max-width:100%;border-radius:6px;border:1px solid #ddd;display:block}
    .imgs figcaption{font-size:11px;color:#7a8295;margin-top:3px}
    a{color:#2557b8;text-decoration:none} a:hover{text-decoration:underline}
    .note{font-size:12px;color:#7a8295}
    .kpis{display:flex;flex-wrap:wrap;gap:12px;margin:10px 0 4px}
    .kpi{background:#f7f9fd;border:1px solid #e1e5ee;border-radius:8px;padding:10px 14px;min-width:150px}
    .kpi .v{font-size:18px;font-weight:600;color:#1F3864} .kpi .l{font-size:11px;color:#6b7488;text-transform:uppercase;letter-spacing:.3px}
    .fit{display:flex;align-items:center;gap:18px;border-radius:10px;padding:16px 20px;margin:4px 0 14px;color:#fff}
    .fit .score{font-size:40px;font-weight:700;line-height:1}
    .fit .meta{font-size:13px;opacity:.95}
    .lane-HOT{background:#c0392b} .lane-WARM{background:#e08e0b} .lane-LATER{background:#5b6b8c} .lane-DEAD{background:#5a5a5a}
    .badge{display:inline-block;font-weight:700;letter-spacing:.5px;font-size:13px;padding:3px 10px;border-radius:6px;background:rgba(255,255,255,.22)}
    .bar{background:#eef1f7;border-radius:5px;height:14px;position:relative;min-width:120px}
    .bar>span{position:absolute;left:0;top:0;bottom:0;background:#3a5a99;border-radius:5px}
    .dq{background:#fdecea;border:1px solid #f5b7b1;color:#922b21;border-radius:8px;padding:8px 12px;margin:6px 0;font-size:13px}
    .acq{background:#eafaf1;border:1px solid #a9dfbf;color:#1e8449;border-radius:8px;padding:8px 12px;margin:6px 0;font-size:13px}
    .flag-red{background:#fdecea;border:1px solid #e6776b;color:#922b21;border-radius:8px;padding:8px 12px;margin:6px 0;font-size:13px}
    .flag-yellow{background:#fef9e7;border:1px solid #f4d03f;color:#9a7d0a;border-radius:8px;padding:8px 12px;margin:6px 0;font-size:13px}
    .headshot{float:right;margin:0 0 10px 16px;text-align:center}
    .headshot img{height:150px;width:150px;object-fit:cover;border-radius:10px;border:1px solid #ddd;display:block}
    .headshot figcaption{font-size:11px;color:#7a8295;margin-top:4px;max-width:160px}
    .dots{letter-spacing:2px;color:#1F3864}
    .msblock{display:flex;flex-direction:column;gap:1px;background:#f0f3fa;border-left:4px solid #1F3864;border-radius:6px;padding:8px 14px;margin:8px 0 4px}
    .msblock .l{font-size:11px;color:#6b7488;text-transform:uppercase;letter-spacing:.3px}
    .msblock .v{font-size:15px;font-weight:600;color:#1F3864}
    """
    sig = viability_signals(prof)

    def money(v):
        return f"${v:,}" if isinstance(v, int) else "—"

    def avg_rating():
        rs = [r.rating for r in prof.reviews if r.rating is not None]
        return round(sum(rs) / len(rs), 1) if rs else None

    P = [f"<!doctype html><html><head><meta charset='utf-8'>"
         f"<title>{_esc(prof.full_name)} — NPI {_esc(prof.npi)}</title>"
         f"<style>{css}</style></head><body><div class='wrap'>"]

    # Lead photo = the main doctor's headshot, shown at the top with the blocks. Prefer a
    # photo whose label matches the doctor's surname; otherwise fall back to the first one.
    lead_photo = None
    ln = (prof.last_name or "").lower()
    if ln:
        lead_photo = next(((lbl, src) for lbl, src in prof.provider_photos
                           if ln in (lbl or "").lower()), None)
    if lead_photo is None and prof.provider_photos:
        lead_photo = prof.provider_photos[0]

    P.append("<div class='card'>")
    if lead_photo:
        P.append(f"<figure class='headshot'><img src='{lead_photo[1]}' alt='{_esc(lead_photo[0])}'>"
                 f"<figcaption>{_esc(lead_photo[0])}</figcaption></figure>")
    _hdr = (f"{prof.full_name} {prof.credential}").strip() or f"NPI {prof.npi}"
    P.append(f"<h1>{_esc(_hdr)}</h1>")
    if prof.display_practice_name:
        P.append(f"<p class='sub'><b>Practice:</b> {_esc(prof.display_practice_name)}</p>")
    P.append(f"<p class='sub'>NPI {_esc(prof.npi)} · {_esc(prof.primary_taxonomy)} · "
             f"report generated {dt.datetime.now():%Y-%m-%d %H:%M}</p>")
    if prof.practice_type:
        P.append(f"<p class='sub'><b>Practice type:</b> {_esc(prof.practice_type)}"
                 f"{(' — ' + _esc(prof.practice_summary)) if prof.practice_summary else ''}</p>")
    if prof.website:
        P.append(f"<p class='sub'><b>Website:</b> "
                 f"<a href='{_esc(prof.website)}' target='_blank'>{_esc(prof.website)}</a></p>")
    if prof.medical_school:
        yr = f" · class of {prof.graduation_year}" if prof.graduation_year else ""
        srcs = f" ({_esc(prof.medical_school_source)})" if prof.medical_school_source else ""
        P.append(f"<div class='msblock'><span class='l'>Medical school</span>"
                 f"<span class='v'>{_esc(prof.medical_school)}{yr}</span>"
                 f"<span class='note'>{srcs}</span></div>")
    if prof.residency:
        rsrc = f" ({_esc(prof.residency_source)})" if prof.residency_source else ""
        P.append(f"<div class='msblock'><span class='l'>Residency</span>"
                 f"<span class='v'>{_esc(prof.residency)}</span>"
                 f"<span class='note'>{rsrc}</span></div>")

    ar = avg_rating()
    grating = next((r for r in prof.reviews if r.source == "Google" and r.rating is not None), None)
    kpis = [("Years in practice*", prof.years_in_practice_proxy or "—"),
            ("Est. age", prof.estimated_age if prof.estimated_age is not None else "—"),
            ("Offices", sig["office_count"]),
            ("Area median HH income", money(sig["area_median_household_income"])),
            ("Median home value", money(sig["area_median_home_value"])),
            ("Medicare beneficiaries", prof.medicare_beneficiaries if prof.medicare_beneficiaries is not None else "—"),
            ("Google rating", f"{grating.rating} ★ ({grating.review_count})" if grating else "—")]
    P.append("<div class='kpis'>")
    for label, val in kpis:
        P.append(f"<div class='kpi'><div class='v'>{_esc(val)}</div><div class='l'>{_esc(label)}</div></div>")
    P.append("</div>")

    # ---- Concierge Fit ----
    if prof.fit_score is not None:
        lane = prof.fit_lane or "DEAD"
        P.append(f"<div class='fit lane-{_esc(lane)}'>")
        P.append(f"<div><div class='score'>{prof.fit_score}</div>"
                 f"<div class='meta'>Concierge Fit / 100</div></div>")
        P.append(f"<div><span class='badge'>{_esc(lane)}</span>"
                 f"<div class='meta'>Specialty: {_esc(prof.specialty_fit_tier or 'unknown')} · "
                 f"Employment: {_esc(prof.employment_status or 'unknown')} · "
                 f"Group size: {_esc(prof.group_size if prof.group_size is not None else '—')}</div></div>")
        P.append("</div>")
        if prof.acquisition_target:
            P.append(f"<div class='acq'><b>★ Acquisition target:</b> {_esc(prof.acquisition_reason)}</div>")
        for level, text in prof.flags:
            cls = "flag-red" if level == "red" else "flag-yellow"
            label = "RED FLAG" if level == "red" else "YELLOW FLAG"
            P.append(f"<div class='{cls}'><b>{label}:</b> {_esc(text)}</div>")
        if prof.disqualifiers:
            for d in prof.disqualifiers:
                P.append(f"<div class='dq'><b>Auto-disqualified:</b> {_esc(d)}</div>")
        if prof.fit_components:
            P.append("<table><tr><th class='key'>Fit component</th><th>Weight</th>"
                     "<th>Subscore</th><th>Contribution</th></tr>")
            for label, w, sub, wtd in prof.fit_components:
                if sub is None:
                    bar = "<span class='note'>no data</span>"
                    subtxt = "—"
                else:
                    bar = (f"<div class='bar'><span style='width:{max(0,min(100,sub))}%'></span></div>")
                    subtxt = f"{sub:g}"
                P.append(f"<tr><th class='key'>{_esc(label)}</th><td>{w}%</td>"
                         f"<td>{bar} {subtxt}</td><td>{('+' + format(wtd,'g')) if wtd is not None else '—'}</td></tr>")
            P.append("</table>")
        P.append("<p class='note'>Fit Score blends affluence/Mosaic, independence (group size), "
                 "Medicare volume, career stage, and specialty; weights renormalize over components "
                 "with data. Hospital/health-system employment, a non-target primary specialty, or a "
                 "deactivated NPI force DEAD regardless of score.</p>")


    P.append("<h2>Identity &amp; registration</h2><table>")
    age_display = ""
    if prof.estimated_age is not None:
        age_display = f"{prof.estimated_age}" + (f" — {prof.age_basis}" if prof.age_basis else "")
    elif prof.age_basis:
        age_display = prof.age_basis  # explains why no estimate was made
    for k, v in [("Estimated age", age_display),
                 ("Enumeration date", prof.enumeration_date), ("Last NPPES update", prof.last_updated),
                 ("Gender", prof.gender), ("Sole proprietor", prof.sole_proprietor),
                 ("Organization", prof.organization_name), ("Licenses", ", ".join(prof.licenses))]:
        if v:
            P.append(f"<tr><th class='key'>{_esc(k)}</th><td>{_esc(v)}</td></tr>")
    P.append("</table>")

    P.append("<h2>Specialties</h2><table>")
    if prof.primary_taxonomy:
        P.append(f"<tr><th class='key'>Primary (NPPES taxonomy)</th><td>{_esc(prof.primary_taxonomy)}</td></tr>")
    if prof.additional_specializations:
        P.append("<tr><th class='key'>Additional (NPPES taxonomies)</th><td>"
                 + "".join(f"<span class='pill'>{_esc(s)}</span>" for s in prof.additional_specializations)
                 + "</td></tr>")
    if prof.dac_primary_specialty:
        P.append(f"<tr><th class='key'>Medicare primary specialty</th><td>{_esc(prof.dac_primary_specialty)}</td></tr>")
    if prof.dac_secondary_specialties:
        P.append("<tr><th class='key'>Medicare secondary specialties</th><td>"
                 + "".join(f"<span class='pill'>{_esc(s)}</span>" for s in prof.dac_secondary_specialties)
                 + "</td></tr>")
    P.append("</table>")

    P.append("<h2>Conversion signals</h2><table>")
    emp = prof.employment_status or "unknown"
    emp_txt = emp + (f" — {prof.employment_evidence}" if prof.employment_evidence else "")
    panel_txt = (f"~{prof.panel_estimate:,} (rough: Medicare benes ÷ 0.30; varies widely)"
                 if prof.panel_estimate else "—")
    ins = {True: "Yes", False: "No / cash-pay", None: "—"}[prof.accepts_insurance]
    if prof.midlevel_roster_known:
        from collections import Counter
        by = Counter(t for _, t in prof.midlevels)
        mid_txt = (f"{prof.midlevel_count} ("
                   + ", ".join(f"{n} {t}" for t, n in by.most_common()) + ")"
                   ) if prof.midlevel_count else "0 (none listed)"
    else:
        mid_txt = "not available"
    tenure_txt = (prof.tenure_at_location
                  + (f" ({prof.tenure_source})" if prof.tenure_source else "")
                  if prof.tenure_at_location else "")
    for k, v in [("Employment status", emp_txt),
                 ("Group size (CMS)", prof.group_size if prof.group_size is not None else "—"),
                 ("Billing organization (CMS)", prof.dac_org_name),
                 ("Current employer (web-verified)", prof.web_current_employer),
                 ("Tenure at current location", tenure_txt),
                 ("Career stage", prof.career_stage),
                 ("Specialty fit tier", prof.specialty_fit_tier),
                 ("Mid-level providers (NP/PA)", mid_txt),
                 ("Medicare beneficiaries", prof.medicare_beneficiaries if prof.medicare_beneficiaries is not None else "—"),
                 ("Est. total panel (informational)", panel_txt),
                 ("Accepts insurance (website)", ins)]:
        if v not in ("", None):
            P.append(f"<tr><th class='key'>{_esc(k)}</th><td>{_esc(v)}</td></tr>")
    P.append("</table>")
    if prof.midlevels:
        P.append("<p class='note'>Mid-levels: "
                 + "; ".join(f"{_esc(n)} ({_esc(t)})" for n, t in prof.midlevels) + "</p>")

    if (prof.competitor_count is not None or prof.nearest_affiliate_mi is not None):
        P.append("<h2>Market context</h2><table>")
        if prof.competitor_count is not None:
            nearest = ("; ".join(f"{n} ({d} mi)" for n, d in prof.competitors)
                       if prof.competitors else "none found")
            P.append(f"<tr><th class='key'>Concierge competitors within {COMPETITION_RADIUS_MI} mi</th>"
                     f"<td><b>{prof.competitor_count}</b> (DPC excluded) — {_esc(nearest)}</td></tr>")
        if prof.nearest_affiliate_mi is not None:
            P.append(f"<tr><th class='key'>Nearest affiliate</th>"
                     f"<td>{_esc(prof.nearest_affiliate_name)} — {prof.nearest_affiliate_mi} mi"
                     f" ({prof.affiliates_nearby} within {AFFILIATE_NEARBY_MI} mi)</td></tr>")
        P.append("</table>")
        P.append("<p class='note'>Market context is informational and not folded into the Fit "
                 "Score. Fewer nearby concierge competitors and a close affiliate (warm intro) "
                 "are favorable.</p>")

    if prof.medicare_beneficiaries is not None or prof.medicare_total_payment is not None:
        yr = f" ({prof.medicare_data_year})" if prof.medicare_data_year else ""
        P.append(f"<h2>Medicare profile{yr}</h2><table>")
        for k, v in [("Provider type", prof.medicare_provider_type),
                     ("Total beneficiaries", prof.medicare_beneficiaries),
                     ("Total services", prof.medicare_total_services),
                     ("Total Medicare payment", money(int(prof.medicare_total_payment)) if prof.medicare_total_payment else None),
                     ("Beneficiaries (medical)", prof.medicare_medical_beneficiaries),
                     ("Beneficiaries (drug)", prof.medicare_drug_beneficiaries)]:
            if v not in (None, ""):
                P.append(f"<tr><th class='key'>{_esc(k)}</th><td>{_esc(v)}</td></tr>")
        P.append("</table>")

    P.append("<h2>Office locations, area wealth &amp; imagery</h2>")
    for a in prof.addresses:
        if not a.is_office:
            continue
        P.append("<div class='office'>")
        P.append(f"<h3>{_esc(a.one_line())}{(' · ' + _esc(a.phone)) if a.phone else ''}</h3>")
        if getattr(a, "source", ""):
            P.append(f"<p class='note' style='margin:0 0 6px'>Source: {_esc(a.source)}</p>")
        P.append("<table>"
                 f"<tr><th class='key'>Area median household income</th><td>{money(a.area_median_household_income)}</td></tr>"
                 f"<tr><th class='key'>Area per-capita income</th><td>{money(a.area_per_capita_income)}</td></tr>"
                 f"<tr><th class='key'>Area median home value</th><td>{money(a.area_median_home_value)}</td></tr>"
                 "</table>")
        if a.mosaic_score is not None:
            hv = a.mosaic_high_value_share
            P.append(f"<p class='note' style='margin:8px 0 2px'><b>Mosaic Score (ZIP {_esc(a.zip5)}):</b> "
                     f"{a.mosaic_score:g}/100 · high-value segment share "
                     f"{hv:g}% · {a.mosaic_classified_pop:,} classified residents</p>")
            P.append("<table><tr><th>Mosaic</th><th class='key'>Segment name</th><th>Tier</th>"
                     "<th>Share of ZIP</th><th>Strength</th></tr>")
            for code, name, seg, cnt, share, lh in a.mosaic_top5:
                dots, word = _lh_indicator(lh)
                P.append(f"<tr><td>{_esc(code)}</td><td>{_esc(str(name).title())}</td>"
                         f"<td>{_esc(seg)}</td><td>{share:g}%</td>"
                         f"<td><span class='dots' title='relative strength'>{dots}</span> "
                         f"<span class='note'>{word}</span></td></tr>")
            P.append("</table>")
        imgs = []
        for u in a.photo_data_uris:
            imgs.append((u, "Office photo (Google)"))
        if a.street_view_data_uri:
            imgs.append((a.street_view_data_uri, "Street View"))
        if a.satellite_data_uri:
            imgs.append((a.satellite_data_uri, "Aerial view"))
        if a.map_data_uri:
            imgs.append((a.map_data_uri, "Location map"))
        if imgs:
            P.append("<div class='imgs'>")
            for src, cap in imgs:
                P.append(f"<figure><a href='{src}' target='_blank'>"
                         f"<img src='{src}' alt='{_esc(cap)}'></a>"
                         f"<figcaption>{_esc(cap)}</figcaption></figure>")
            P.append("</div>")
        elif a.map_link:
            P.append(f"<p class='note'>Map: <a href='{_esc(a.map_link)}' target='_blank'>view location on OpenStreetMap</a></p>")
        P.append("</div>")
    superseded = [a for a in prof.addresses if a.superseded]
    mailing = [a for a in prof.addresses if not a.is_office and not a.superseded]
    if mailing:
        P.append("<p class='note'>Mailing: " + "; ".join(_esc(a.one_line()) for a in mailing) + "</p>")
    if superseded:
        P.append("<p class='note'><b>NPPES-registered address (superseded — may be outdated):</b> "
                 + "; ".join(_esc(a.one_line()) for a in superseded)
                 + ". The office address above was taken from a current source (practice website "
                 "or Google listing).</p>")

    P.append("<h2>Web presence</h2><table>")
    if prof.website:
        P.append(f"<tr><th class='key'>Website</th><td><a href='{_esc(prof.website)}' target='_blank'>{_esc(prof.website)}</a></td></tr>")
    if prof.website_services:
        P.append("<tr><th class='key'>Services offered</th><td>"
                 + "".join(f"<span class='pill'>{_esc(s)}</span>" for s in prof.website_services) + "</td></tr>")
    if prof.website_insurances:
        P.append("<tr><th class='key'>Insurances accepted</th><td>"
                 + "".join(f"<span class='pill'>{_esc(s)}</span>" for s in prof.website_insurances) + "</td></tr>")
    P.append("</table>")

    if prof.interior_photos:
        P.append("<h2>Practice interior photos</h2><div class='imgs'>")
        for src in prof.interior_photos:
            P.append(f"<figure><a href='{src}' target='_blank'>"
                     f"<img src='{src}' alt='Practice interior'></a></figure>")
        P.append("</div>")
        P.append("<p class='note'>Interior photos captured from the practice website"
                 + (" and confirmed by image review to show the office interior."
                    if ANTHROPIC_API_KEY else "; classification is best-effort.") + "</p>")

    P.append("<h2>Reviews &amp; ratings</h2>")
    rated = [r for r in prof.reviews if r.rating is not None]
    if rated:
        P.append("<table><tr><th class='key'>Source</th><th>Rating</th><th>Reviews</th></tr>")
        for r in rated:
            cnt = r.review_count if r.review_count is not None else "—"
            P.append(f"<tr><th class='key'>{_esc(r.source)}</th>"
                     f"<td>{_esc(r.rating)} ★</td><td>{_esc(cnt)}</td></tr>")
        if ar is not None and len(rated) > 1:
            P.append(f"<tr><th class='key'>Average (all sources)</th>"
                     f"<td><b>{_esc(ar)} ★</b></td><td>—</td></tr>")
        P.append("</table>")
    else:
        P.append("<p class='note'>No aggregate rating could be auto-extracted from public sources.</p>")

    if prof.external_profiles:
        P.append("<h2>Physician directory profiles</h2><table>"
                 "<tr><th class='key'>Directory</th><th>Rating</th><th>Profile</th></tr>")
        for p in prof.external_profiles:
            if p.rating is not None:
                cnt = f" ({p.review_count})" if p.review_count is not None else ""
                rating_cell = f"{_esc(p.rating)} ★{_esc(cnt)}"
            else:
                rating_cell = _esc(p.note) if p.note else "—"
            P.append(f"<tr><th class='key'>{_esc(p.source)}</th><td>{rating_cell}</td>"
                     f"<td><a href='{_esc(p.url)}' target='_blank'>open profile</a></td></tr>")
        P.append("</table>")
        P.append("<p class='note'>Ratings, where shown, are folded into the blended average above. "
                 "Directory ratings are auto-extracted and should be verified on the source page.</p>")

    P.append("<h2>Board certification &amp; recognition</h2>")
    if prof.board_certifications or prof.recognitions:
        P.append("<table>")
        if prof.board_certifications:
            P.append("<tr><th class='key'>Board certification (found)</th><td>"
                     + "".join(f"<span class='pill'>{_esc(c)}</span>" for c in prof.board_certifications)
                     + "</td></tr>")
        if prof.recognitions:
            P.append("<tr><th class='key'>Recognition / awards (found)</th><td>"
                     + "".join(f"<span class='pill'>{_esc(r)}</span>" for r in prof.recognitions)
                     + "</td></tr>")
        P.append("</table>")
        P.append("<p class='note'>Auto-extracted from the provider's website / public pages; verify.</p>")
    else:
        P.append("<p class='note'>No board-certification or recognition text was found on the "
                 "provider's website / public pages.</p>")

    if prof.personal_interests:
        P.append("<h2>Personal interests &amp; rapport</h2>"
                 "<p class='note'>Publicly shared, non-clinical details for outreach — verify "
                 "before referencing.</p><ul>")
        for it in prof.personal_interests:
            src = (it.get("source") or "").strip()
            if src.startswith("http"):
                tail = f" <span class='note'>(<a href='{_esc(src)}' target='_blank'>source</a>)</span>"
            elif src:
                tail = f" <span class='note'>({_esc(src)})</span>"
            else:
                tail = ""
            P.append(f"<li>{_esc(it.get('text',''))}{tail}</li>")
        P.append("</ul>")

    if prof.notes:
        P.append("<h2>Notes &amp; data gaps</h2><ul class='note'>")
        for n in prof.notes:
            P.append(f"<li>{_esc(n)}</li>")
        P.append("</ul>")
    P.append(f"<p class='note'>Sources used: {_esc(', '.join(prof.sources_used) or 'none')}. "
             f"*Years in practice is a proxy (time since NPI enumeration).</p>")
    P.append("</div></div></body></html>")

    path = os.path.join(outdir, f"{prof.npi}.html")
    with open(path, "w", encoding="utf-8") as f:
        f.write("".join(P))
    return path


def open_in_chrome(path: str) -> None:
    """Open the report in Chrome (falls back to the default browser)."""
    uri = Path(path).resolve().as_uri()
    for exe in (r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"):
        if os.path.exists(exe):
            webbrowser.register("chrome", None, webbrowser.BackgroundBrowser(exe))
            webbrowser.get("chrome").open(uri)
            return
    webbrowser.open(uri)


# ----------------------------------------------------------------------------
# Batch mode
# ----------------------------------------------------------------------------
_LANE_FILL = {"HOT": "C0392B", "WARM": "E08E0B", "LATER": "5B6B8C", "DEAD": "7A7A7A"}
_LANE_ORDER = {"HOT": 0, "WARM": 1, "LATER": 2, "DEAD": 3, "": 4}


def _read_npis(csv_path: str) -> list[str]:
    import csv
    npis = []
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(2048)
        f.seek(0)
        has_header = csv.Sniffer().has_header(sample) if sample.strip() else False
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return npis
    col = 0
    start = 0
    if has_header:
        header = [h.strip().lower() for h in rows[0]]
        col = next((i for i, h in enumerate(header) if h in ("npi", "npi_number", "npi #", "npi#")), 0)
        start = 1
    for r in rows[start:]:
        if col < len(r):
            digits = re.sub(r"\D", "", str(r[col]))
            if len(digits) == 10:
                npis.append(digits)
    # de-dup, preserve order
    seen, out = set(), []
    for n in npis:
        if n not in seen:
            seen.add(n)
            out.append(n)
    return out


def write_batch_xlsx(profiles: list[ProviderProfile], out_path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"
    cols = ["Fit Score", "Lane", "NPI", "Doctor", "Practice", "Primary specialty",
            "Specialty tier", "Employment", "Group size", "Est. age", "Career stage",
            "Midlevels", "Midlevel types",
            "Medicare benes", "Area median income", "Mosaic score", "High-value share %",
            "Competitors (10mi)", "Nearest affiliate (mi)",
            "Offices", "City", "State", "Website", "Disqualifiers", "Personal interests",
            "Report file"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", start_color="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def rank_key(p):
        return (_LANE_ORDER.get(p.fit_lane, 4), -(p.fit_score or 0))

    from collections import Counter
    for p in sorted(profiles, key=rank_key):
        loc = _primary_office(p)
        if p.midlevel_roster_known:
            mid_n = p.midlevel_count
            by = Counter(t for _, t in p.midlevels)
            mid_types = ", ".join(f"{n} {t}" for t, n in by.most_common())
        else:
            mid_n, mid_types = "n/a", ""
        ws.append([
            p.fit_score if p.fit_score is not None else "",
            p.fit_lane, p.npi, p.full_name, p.display_practice_name,
            p.primary_taxonomy or p.dac_primary_specialty, p.specialty_fit_tier,
            p.employment_status, p.group_size if p.group_size is not None else "",
            p.estimated_age if p.estimated_age is not None else "", p.career_stage,
            mid_n, mid_types,
            p.medicare_beneficiaries if p.medicare_beneficiaries is not None else "",
            (loc.area_median_household_income if loc and isinstance(loc.area_median_household_income, int) else ""),
            (loc.mosaic_score if loc and loc.mosaic_score is not None else ""),
            (loc.mosaic_high_value_share if loc and loc.mosaic_high_value_share is not None else ""),
            (p.competitor_count if p.competitor_count is not None else ""),
            (p.nearest_affiliate_mi if p.nearest_affiliate_mi is not None else ""),
            len([a for a in p.addresses if a.is_office]),
            loc.city if loc else "", loc.state if loc else "", p.website,
            "; ".join(p.disqualifiers),
            "; ".join(i["text"] for i in p.personal_interests),
            os.path.basename(p.npi + ".html"),
        ])
        lane_cell = ws.cell(row=ws.max_row, column=2)
        fill = _LANE_FILL.get(p.fit_lane)
        if fill:
            lane_cell.fill = PatternFill("solid", start_color=fill)
            lane_cell.font = Font(bold=True, color="FFFFFF", name="Arial")
    for col_cells in ws.iter_cols(min_row=2):
        for cell in col_cells:
            if cell.font is None or cell.font.name != "Arial":
                cell.font = Font(name="Arial")
    widths = [9, 8, 12, 24, 26, 22, 11, 14, 9, 8, 22, 10, 22, 13, 16, 11, 14, 16, 18, 7, 16, 6, 34, 40, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    wb.save(out_path)
    return out_path


def run_batch(csv_path: str, *, do_google=True, do_scrape=True, do_aggregators=True,
              do_photos=True, do_web=True, do_interests=True, cms_uuid="", out_xlsx="",
              push_monday=False, monday_board="", monday_group="") -> None:
    if not _HAVE_OPENPYXL:
        print("[error] Batch mode needs openpyxl (pip install openpyxl).", file=sys.stderr)
        return
    npis = _read_npis(csv_path)
    if not npis:
        print(f"[error] No valid 10-digit NPIs found in {csv_path} "
              "(expects a column named 'NPI').", file=sys.stderr)
        return
    outdir = tempfile.mkdtemp(prefix="npi_scout_batch_")
    print(f"[batch] {len(npis)} NPIs -> reports in {outdir}", file=sys.stderr)
    profiles = []
    for i, npi in enumerate(npis, 1):
        print(f"\n[batch {i}/{len(npis)}] {npi}", file=sys.stderr)
        try:
            prof = research_npi(npi, cms_uuid=cms_uuid, do_google=do_google,
                                do_scrape=do_scrape, do_aggregators=do_aggregators,
                                do_photos=do_photos, do_web=do_web, do_interests=do_interests)
        except Exception as e:  # noqa: BLE001
            print(f"    [warn] {npi} failed: {e}", file=sys.stderr)
            prof = ProviderProfile(npi=npi)
            prof.notes.append(f"Processing error: {e}")
        if prof.full_name:
            write_html(prof, outdir)
        profiles.append(prof)
    out_xlsx = out_xlsx or os.path.join(outdir, "prospects_ranked.xlsx")
    write_batch_xlsx(profiles, out_xlsx)
    ranked = sorted((p for p in profiles if p.fit_score is not None),
                    key=lambda p: (_LANE_ORDER.get(p.fit_lane, 4), -(p.fit_score or 0)))
    print(f"\n[batch] Wrote ranked workbook: {out_xlsx}", file=sys.stderr)
    print(f"[batch] Per-doctor reports: {outdir}", file=sys.stderr)
    hot = [p for p in profiles if p.fit_lane == "HOT"]
    warm = [p for p in profiles if p.fit_lane == "WARM"]
    print(f"[batch] HOT: {len(hot)} · WARM: {len(warm)} · "
          f"DEAD: {len([p for p in profiles if p.fit_lane=='DEAD'])}", file=sys.stderr)
    if push_monday:
        print("[batch] Pushing to monday.com ...", file=sys.stderr)
        push_to_monday(profiles, board_id=monday_board, group_id=monday_group)
    try:
        open_in_chrome(out_xlsx)
    except Exception:  # noqa: BLE001
        pass


# ----------------------------------------------------------------------------
# monday.com push
# ----------------------------------------------------------------------------
def _monday_post(query: str, variables: dict | None = None):
    if not MONDAY_API_TOKEN:
        return None
    headers = {"Authorization": MONDAY_API_TOKEN, "Content-Type": "application/json",
               "API-Version": MONDAY_API_VERSION}
    body = {"query": query, "variables": variables or {}}
    resp = _post(MONDAY_API_URL, json_body=body, headers=headers)
    if resp and resp.get("errors"):
        msg = "; ".join(e.get("message", "") for e in resp["errors"])
        print(f"    [warn] monday API error: {msg}", file=sys.stderr)
        return None
    return resp


def monday_board_columns(board_id: str) -> dict:
    """Return {title_lower: (column_id, column_type)} for the board."""
    q = "query($b:[ID!]){boards(ids:$b){columns{id title type}}}"
    resp = _monday_post(q, {"b": [str(board_id)]})
    cols = {}
    try:
        for c in resp["data"]["boards"][0]["columns"]:
            cols[str(c["title"]).strip().lower()] = (c["id"], c["type"])
    except (TypeError, KeyError, IndexError):
        pass
    return cols


def _monday_encode(ctype: str, value):
    """Encode a Python value for a monday column of the given type."""
    if value in (None, "", []):
        return None
    t = (ctype or "").lower()
    if t in ("numbers", "numeric"):
        try:
            return str(float(value)).rstrip("0").rstrip(".") if "." in str(value) else str(int(value))
        except (ValueError, TypeError):
            return None
    if t in ("status", "color", "dropdown"):
        return {"label": str(value)}
    if t in ("link",):
        return {"url": str(value), "text": str(value)}
    if t in ("long_text", "long-text", "longtext"):
        return {"text": str(value)[:2000]}
    return str(value)  # text / default


# our field -> (column title to match on the board, getter(prof, primary_office))
def _monday_field_map():
    def inc(loc):
        return loc.area_median_household_income if loc and isinstance(loc.area_median_household_income, int) else None
    return [
        ("Fit Score", lambda p, l: p.fit_score),
        ("Lane", lambda p, l: p.fit_lane or None),
        ("NPI", lambda p, l: p.npi),
        ("Practice", lambda p, l: p.display_practice_name or None),
        ("Primary Specialty", lambda p, l: p.primary_taxonomy or p.dac_primary_specialty or None),
        ("Specialty Tier", lambda p, l: p.specialty_fit_tier or None),
        ("Employment", lambda p, l: p.employment_status or None),
        ("Group Size", lambda p, l: p.group_size),
        ("Est Age", lambda p, l: p.estimated_age),
        ("Career Stage", lambda p, l: p.career_stage or None),
        ("Residency", lambda p, l: p.residency or None),
        ("Tenure", lambda p, l: p.tenure_at_location or None),
        ("Personal Interests", lambda p, l: "; ".join(i["text"] for i in p.personal_interests) or None),
        ("Medicare Beneficiaries", lambda p, l: p.medicare_beneficiaries),
        ("Area Median Income", lambda p, l: inc(l)),
        ("Mosaic Score", lambda p, l: l.mosaic_score if l else None),
        ("High-Value Share", lambda p, l: l.mosaic_high_value_share if l else None),
        ("Competitors 10mi", lambda p, l: p.competitor_count),
        ("Nearest Affiliate Mi", lambda p, l: p.nearest_affiliate_mi),
        ("Offices", lambda p, l: len([a for a in p.addresses if a.is_office]) or None),
        ("City", lambda p, l: l.city if l else None),
        ("State", lambda p, l: l.state if l else None),
        ("Website", lambda p, l: p.website or None),
        ("Disqualifiers", lambda p, l: "; ".join(p.disqualifiers) or None),
    ]


def push_to_monday(profiles: list[ProviderProfile], board_id: str = "",
                   group_id: str = "") -> None:
    board_id = board_id or MONDAY_BOARD_ID
    group_id = group_id or MONDAY_GROUP_ID
    if not MONDAY_API_TOKEN or not board_id:
        print("    [warn] monday push skipped: set MONDAY_API_TOKEN and MONDAY_BOARD_ID "
              "(or pass --monday-board).", file=sys.stderr)
        return
    cols = monday_board_columns(board_id)
    if not cols:
        print("    [warn] monday push: could not read board columns (check token/board id).",
              file=sys.stderr)
        return
    # title aliases so minor naming differences still match
    aliases = {"est age": "est. age", "high-value share": "high value share"}
    fmap = _monday_field_map()
    matched = [t for t, _ in fmap if t.lower() in cols or aliases.get(t.lower(), "") in cols]
    print(f"    [monday] matched {len(matched)} of {len(fmap)} fields to board columns "
          f"by title.", file=sys.stderr)

    mut = ("mutation($b:ID!,$g:String,$n:String!,$v:JSON!){"
           "create_item(board_id:$b,group_id:$g,item_name:$n,column_values:$v,"
           "create_labels_if_missing:true){id}}")
    created = 0
    for p in profiles:
        if not p.full_name:
            continue
        loc = _primary_office(p)
        vals = {}
        for title, getter in fmap:
            key = title.lower()
            col = cols.get(key) or cols.get(aliases.get(key, ""))
            if not col:
                continue
            enc = _monday_encode(col[1], getter(p, loc))
            if enc is not None:
                vals[col[0]] = enc
        v = {"b": str(board_id), "g": group_id or None,
             "n": p.full_name[:255], "v": json.dumps(vals)}
        resp = _monday_post(mut, v)
        if resp and resp.get("data", {}).get("create_item", {}).get("id"):
            created += 1
        time.sleep(0.3)  # stay under complexity limits
    print(f"    [monday] created {created} item(s) on board {board_id}.", file=sys.stderr)


def monday_inspect(board_id: str = "") -> None:
    board_id = board_id or MONDAY_BOARD_ID
    if not MONDAY_API_TOKEN or not board_id:
        print("Set MONDAY_API_TOKEN and MONDAY_BOARD_ID (or --monday-board) first.", file=sys.stderr)
        return
    cols = monday_board_columns(board_id)
    if not cols:
        print("No columns returned — check the token and board id.", file=sys.stderr)
        return
    print(f"Board {board_id} columns (create these titles to receive data):", file=sys.stderr)
    wanted = {t.lower() for t, _ in _monday_field_map()}
    for title, (cid, ctype) in sorted(cols.items()):
        mark = "  <- matched" if title in wanted else ""
        print(f"   '{title}'  type={ctype}  id={cid}{mark}", file=sys.stderr)
    have = {t for t, _ in _monday_field_map() if t.lower() in cols}
    missing = [t for t, _ in _monday_field_map() if t.lower() not in cols]
    print(f"\nMatched {len(have)} fields. Add columns titled: {', '.join(missing)}", file=sys.stderr)


# ----------------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------------
def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Research physicians by NPI and score concierge-medicine fit. "
                    "Single NPI opens an HTML report; --batch ranks a CSV of NPIs to xlsx.")
    ap.add_argument("npi", nargs="?", help="A 10-digit NPI. If omitted (and no --batch), you'll be prompted.")
    ap.add_argument("--batch", default="", metavar="CSV",
                    help="Path to a CSV with an 'NPI' column; ranks all of them to an xlsx.")
    ap.add_argument("--out", default="", metavar="XLSX",
                    help="Output path for the ranked workbook in batch mode.")
    ap.add_argument("--json", action="store_true",
                    help="Also write the raw JSON record alongside the report (single mode).")
    ap.add_argument("--cms-dataset-uuid", default="",
                    help="Override the CMS 'by Provider' dataset UUID if auto-discovery fails.")
    ap.add_argument("--no-google", action="store_true", help="Skip Google Places enrichment.")
    ap.add_argument("--no-scrape", action="store_true", help="Skip website/review scraping.")
    ap.add_argument("--no-aggregators", action="store_true",
                    help="Skip physician-directory lookups (Healthgrades/U.S. News/Vitals/WebMD/Zocdoc).")
    ap.add_argument("--no-photos", action="store_true",
                    help="Skip capturing interior/provider photos from the practice website.")
    ap.add_argument("--no-web", action="store_true",
                    help="Skip the open web search used to find the doctor's CURRENT practice "
                         "location (falls back to website/Google/NPPES).")
    ap.add_argument("--no-interests", action="store_true",
                    help="Skip the open web search for non-clinical personal interests / rapport "
                         "facts (hobbies, sports, music, volunteering).")
    ap.add_argument("--monday", action="store_true",
                    help="Push result(s) to a monday.com board (needs MONDAY_API_TOKEN + board id).")
    ap.add_argument("--monday-board", default="", help="monday board id (overrides MONDAY_BOARD_ID).")
    ap.add_argument("--monday-group", default="", help="monday group id (optional).")
    ap.add_argument("--monday-inspect", action="store_true",
                    help="List the monday board's columns (to see which titles will be matched) and exit.")
    args = ap.parse_args(argv)

    if args.monday_inspect:
        monday_inspect(args.monday_board)
        return

    if args.batch:
        run_batch(args.batch, do_google=not args.no_google, do_scrape=not args.no_scrape,
                  do_aggregators=not args.no_aggregators, do_photos=not args.no_photos,
                  do_web=not args.no_web, do_interests=not args.no_interests,
                  cms_uuid=args.cms_dataset_uuid,
                  out_xlsx=args.out, push_monday=args.monday,
                  monday_board=args.monday_board, monday_group=args.monday_group)
        return

    npi = args.npi
    if not npi:
        npi = input("Enter the 10-digit NPI number to research: ").strip()

    outdir = tempfile.mkdtemp(prefix="npi_scout_")
    prof = research_npi(npi, cms_uuid=args.cms_dataset_uuid,
                        do_google=not args.no_google, do_scrape=not args.no_scrape,
                        do_aggregators=not args.no_aggregators, do_photos=not args.no_photos,
                        do_web=not args.no_web, do_interests=not args.no_interests)
    html = write_html(prof, outdir)
    if args.json:
        write_json(prof, outdir)
    if args.monday and prof.full_name:
        print("Pushing to monday.com ...")
        push_to_monday([prof], board_id=args.monday_board, group_id=args.monday_group)
    open_in_chrome(html)
    print(f"\nDone. Opened report in Chrome:\n  {html}")
    if not prof.full_name:
        notes = " ".join(prof.notes).lower()
        if "not a valid npi" in notes:
            print("  (That NPI is not valid — check the number.)")
        else:
            print("  (Valid NPI, but no provider name could be found in NPPES, CMS, or on the "
                  "open web. The NPI may be deactivated/withheld. The report shows whatever the "
                  "other sources returned.)")


if __name__ == "__main__":
    main()